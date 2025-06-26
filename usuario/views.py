import datetime
import json
from django.shortcuts import render, redirect
import psycopg2
from .forms import LoginForm
from django.contrib.auth import authenticate, login
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
import openpyxl
from django.db import connection
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger # Importa Paginator
from .forms import ExamenesFilterForm,EgresadosFilterForm,CursadasFilterForm,ResultadoCursaFilterForm,IngresantesFilterForm,DocentesFilterForm,EgresadosxAnioFilterForm
from itertools import groupby
from operator import itemgetter
from datetime import date, datetime # Importa ambos, date y datetime

# Create your views here.


def user_login(request):
    if request.method == 'POST':

         # Se crea una instancia del formulario con los datos que vienen en la solicitud
        form = LoginForm(request.POST)

        if form.is_valid():
            # Se obtiene un diccionario con los datos validados del formulario
            cd = form.cleaned_data

            # Se intenta autenticar al usuario con los datos ingresados
            user =  authenticate(request, username=cd['username'], password=cd['password'])

            # Si el usuario existe y las credenciales son correctas
            if user is not None:
                if user.is_active:

                    # Se inicia sesión para ese usuario
                    login(request, user)
                    # Redirigir a la página de inicio o a la página deseada
                    return HttpResponse('Usuario Logueado Correctamente')
                else:
                    return HttpResponse('Usuario Inactivo')
            else:
                # El usuario no es válido, mostrar un mensaje de error
                return HttpResponse('Usuario o contraseña incorrectos')

    else:

        # Si el método de la solicitud no es POST (ej. GET), se muestra el formulario vacío
        form = LoginForm()

    # Finalmente, se renderiza la plantilla de login con el formulario (vacío o con errores)
    return render(request, 'account/login.html', {'form': form})

@login_required
def dashboard(request):
    # Aquí puedes agregar la lógica para mostrar el panel de control del usuario
    return render(request, 'account/dashboard.html', {'section': 'dashboard'})

@login_required
def inicio_view(request):
    current_year = date.today().year

    # Estructuras para almacenar los datos de cada tipo de grado para Chart.js
    grado_chart_data = {
        'labels': [],
        'data': [],
        'background_colors': [],
        'year': current_year
    }
    pregrado_chart_data = {
        'labels': [],
        'data': [],
        'background_colors': [],
        'year': current_year
    }
    posgrado_chart_data = {
        'labels': [],
        'data': [],
        'background_colors': [],
        'year': current_year
    }

    # Definimos las paletas de colores una vez
    colors_palette_grado = [
        'rgba(255, 99, 132, 0.7)', 'rgba(54, 162, 235, 0.7)', 'rgba(255, 206, 86, 0.7)',
        'rgba(75, 192, 192, 0.7)', 'rgba(153, 102, 255, 0.7)', 'rgba(255, 159, 64, 0.7)',
        'rgba(199, 199, 199, 0.7)', 'rgba(83, 102, 255, 0.7)', 'rgba(201, 203, 207, 0.7)',
        'rgba(255, 0, 0, 0.7)', 'rgba(0, 255, 0, 0.7)', 'rgba(0, 0, 255, 0.7)'
    ]
    colors_palette_pregrado = [
        'rgba(50, 205, 50, 0.7)', 'rgba(0, 128, 128, 0.7)', 'rgba(255, 140, 0, 0.7)',
        'rgba(138, 43, 226, 0.7)', 'rgba(0, 206, 209, 0.7)', 'rgba(255, 69, 0, 0.7)',
        'rgba(123, 104, 238, 0.7)', 'rgba(210, 105, 30, 0.7)', 'rgba(124, 252, 0, 0.7)'
    ]
    colors_palette_posgrado = [
        'rgba(128, 0, 128, 0.7)', 'rgba(0, 0, 255, 0.7)', 'rgba(255, 215, 0, 0.7)',
        'rgba(70, 130, 180, 0.7)', 'rgba(218, 112, 214, 0.7)', 'rgba(255, 99, 71, 0.7)',
        'rgba(60, 179, 113, 0.7)', 'rgba(72, 61, 139, 0.7)', 'rgba(173, 216, 230, 0.7)'
    ]

    try:
        with connection.cursor() as cursor:
            # Definimos las ramas de la consulta UNION ALL
            # Cada rama necesita un placeholder para el anio_academico
            # Y se usará siempre, a diferencia del ejemplo de docentes donde era condicional.
            sql_branch_grado = f"""
                SELECT
                    'Grado' as tipo_grado,
                    EXTRACT(YEAR FROM CURRENT_DATE) as anio,
                    pr.codigo as codigo,
                    pr.nombre as carrera,
                    count(DISTINCT sa.legajo) as cantidad_alumnos -- Usar DISTINCT para legajo si un alumno puede estar en varias comisiones del mismo tipo de propuesta
                FROM
                     negocio.mdp_personas mp
                INNER JOIN negocio.sga_alumnos sa ON mp.persona = sa.persona
                INNER JOIN negocio.sga_propuestas pr ON sa.propuesta = pr.propuesta
                WHERE
                    pr.codigo not IN ('L', 'F', 'E')
                    and pr.propuesta_tipo = 200
                    AND EXISTS (
                        SELECT 1
                        FROM
                            negocio.vw_comisiones c
                        INNER JOIN negocio.sga_insc_cursada ic ON c.comision = ic.comision
                        WHERE
                            ic.alumno = sa.alumno
                            AND c.anio_academico = %s
                            AND ic.estado = 'A'
                    )
                  GROUP BY
                    pr.codigo,
                    pr.nombre
            """

            sql_branch_pregrado = f"""
                SELECT
                    'Pre-Grado' as tipo_grado,
                    EXTRACT(YEAR FROM CURRENT_DATE) as anio,
                    pr.codigo as codigo,
                    pr.nombre as carrera,
                    count(DISTINCT sa.legajo) as cantidad_alumnos
                FROM
                     negocio.mdp_personas mp
                INNER JOIN negocio.sga_alumnos sa ON mp.persona = sa.persona
                INNER JOIN negocio.sga_propuestas pr ON sa.propuesta = pr.propuesta
                WHERE
                     pr.propuesta_tipo = 201
                    AND EXISTS (
                        SELECT 1
                        FROM
                            negocio.vw_comisiones c
                        INNER JOIN negocio.sga_insc_cursada ic ON c.comision = ic.comision
                        WHERE
                            ic.alumno = sa.alumno
                            AND c.anio_academico  = %s
                            AND ic.estado = 'A'
                    )
                  GROUP BY
                    pr.codigo,
                    pr.nombre
            """

            sql_branch_posgrado = f"""
                SELECT
                    'Posgrado' as tipo_grado,
                    EXTRACT(YEAR FROM CURRENT_DATE) as anio,
                    pr.codigo as codigo,
                    pr.nombre as carrera,
                    count(DISTINCT sa.legajo) as cantidad_alumnos
                FROM
                     negocio.mdp_personas mp
                INNER JOIN negocio.sga_alumnos sa ON mp.persona = sa.persona
                INNER JOIN negocio.sga_propuestas pr ON sa.propuesta = pr.propuesta
                WHERE
                     pr.propuesta_tipo = 202
                    AND EXISTS (
                        SELECT 1
                        FROM
                            negocio.vw_comisiones c
                        INNER JOIN negocio.sga_insc_cursada ic ON c.comision = ic.comision
                        WHERE
                            ic.alumno = sa.alumno
                            AND c.anio_academico = %s
                            AND ic.estado = 'A'
                    )
                  GROUP BY
                    pr.codigo,
                    pr.nombre
            """

            # Unir las ramas con UNION ALL
            full_sql_query = f"""
                {sql_branch_grado}
                UNION ALL
                {sql_branch_pregrado}
                UNION ALL
                {sql_branch_posgrado}
                ORDER BY tipo_grado, carrera;
            """

            # Los parámetros para la consulta completa.
            # Cada %s en las ramas requiere un valor en esta lista.
            # Como cada rama usa %s para el año, lo pasamos 3 veces.
            sql_params = [current_year, current_year, current_year]

            cursor.execute(full_sql_query, sql_params)
            results = cursor.fetchall()

            # Procesar los resultados y asignarlos a las estructuras de datos correctas
            grado_count = 0
            pregrado_count = 0
            posgrado_count = 0

            for row in results:
                tipo_grado, anio, codigo, carrera, cantidad_alumnos = row
                label = f"{carrera} ({codigo})"

                if tipo_grado == 'Grado':
                    grado_chart_data['labels'].append(label)
                    grado_chart_data['data'].append(cantidad_alumnos)
                    grado_chart_data['background_colors'].append(colors_palette_grado[grado_count % len(colors_palette_grado)])
                    grado_count += 1
                elif tipo_grado == 'Pre-Grado':
                    pregrado_chart_data['labels'].append(label)
                    pregrado_chart_data['data'].append(cantidad_alumnos)
                    pregrado_chart_data['background_colors'].append(colors_palette_pregrado[pregrado_count % len(colors_palette_pregrado)])
                    pregrado_count += 1
                elif tipo_grado == 'Posgrado':
                    posgrado_chart_data['labels'].append(label)
                    posgrado_chart_data['data'].append(cantidad_alumnos)
                    posgrado_chart_data['background_colors'].append(colors_palette_posgrado[posgrado_count % len(colors_palette_posgrado)])
                    posgrado_count += 1

    except psycopg2.Error as e:
        print(f"Error de base de datos en inicio_view: {e}")
        # En producción, podrías loggear esto y quizás pasar un mensaje de error al template
        # para que el usuario sepa que algo salió mal.
    except Exception as e:
        print(f"Error inesperado en inicio_view: {e}")
        # Lo mismo aquí, para errores inesperados.

    context = {
        'grado_chart_data': grado_chart_data,
        'pregrado_chart_data': pregrado_chart_data,
        'posgrado_chart_data': posgrado_chart_data,
    }

    # Renderiza el template parcial, tal como lo haces actualmente.
    return render(request, 'account/inicio/inicio.html', context)


@login_required
def reportes_view(request):
    # Lógica para obtener reportes
    return render(request, 'account/reportes/reportes.html', {'reportes': []})

@login_required
def estadisticas_view(request):
    # Lógica para obtener estaditicas
    return render(request, 'account/estadisticas/estadisticas.html', {'estadisticas': []})

@login_required
def egresados_report(request):
    form = EgresadosFilterForm(request.GET)
    egresados = []
    egresados_page_obj = None
    report_executed = False

    if 'anio' in request.GET or 'propuesta_ids' in request.GET:
        if form.is_valid():
            anio_filter = form.cleaned_data.get('anio')
            propuesta_ids_filter = form.cleaned_data.get('propuesta_ids')

            if anio_filter and propuesta_ids_filter:
                report_executed = True

                propuesta_ids_str = ','.join(map(str, [int(p_id) for p_id in propuesta_ids_filter]))

                with connection.cursor() as cursor:
                    sql_query = f"""
                                 select vp.apellido_nombres as nombre, n.descripcion as nacionalidad, negocio_pers.anioingreso(a.alumno) as anioingreso,
                                 negocio_pers.get_mail(a.persona) as mail,vp.tipo_nro_documento as dni,  vp.sexo as sexo, a.legajo as legajo, a.propuesta_nombre as carrera,
                                 a.plan_codigo as codigo, sco.fecha_egreso as egreso, sco.promedio as promedio, sco.promedio_sin_aplazos as sin_aplazos
                                 from negocio.vw_alumnos a,
                                     negocio.sga_certificados_otorg sco,
                                    negocio.mdp_nacionalidades n ,
                                negocio.vw_personas vp
                                where a.persona = vp.persona
                                and   a.propuesta in ({propuesta_ids_str})
                                and   a.alumno = sco.alumno
                                and   vp.persona = sco.persona
                                and   vp.nacionalidad = n.nacionalidad
                                and   EXTRACT(YEAR FROM sco.fecha_egreso) = %s
                                order by vp.apellido_nombres
                        """
                    cursor.execute(sql_query, [anio_filter])

                    columns = [col[0] for col in cursor.description]
                    all_egresados_data = cursor.fetchall()

                    egresados_list = []
                    for row in all_egresados_data:
                        row_dict = dict(zip(columns, row))
                        egresados_list.append(row_dict)

                # El paginador y el formato de datos deben ocurrir después de obtener todos los datos,
                # pero pueden estar fuera del bloque del cursor ya que all_egresados_data ya está en memoria.
                paginator = Paginator(egresados_list, 25)
                page_number = request.GET.get('page')

                try:
                    egresados_page_obj = paginator.page(page_number)
                except PageNotAnInteger:
                    egresados_page_obj = paginator.page(1)
                except EmptyPage:
                    egresados_page_obj = paginator.page(paginator.num_pages)

                # Poblar la lista 'egresados' directamente desde egresados_page_obj.
                # Esto asegura que los datos estén correctamente estructurados para la plantilla
                # y evita posibles problemas de re-iteración con el cursor.
                for egresado_dict in egresados_page_obj:
                    egresados.append({
                        'nombre': egresado_dict['nombre'],
                        'nacionalidad': egresado_dict['nacionalidad'],
                        'ingreso': egresado_dict['anioingreso'],
                        'mail': egresado_dict['mail'],
                        'dni': egresado_dict['dni'],
                        'sexo': egresado_dict['sexo'],
                        'legajo': egresado_dict['legajo'],
                        'carrera': egresado_dict['carrera'],
                        'codigo': egresado_dict['codigo'],
                        'egreso': egresado_dict['egreso'],
                        'promedio': egresado_dict['promedio'],
                        'sin_aplazos': egresado_dict['sin_aplazos'],
                    })

    context = {
        'form': form,
        'egresados': egresados,
        'egresados_page_obj': egresados_page_obj,
        'report_title': 'Reporte de Egresados',
        'report_executed': report_executed,
    }

    return render(request, 'account/reportes/egresados/egresados_report.html', context)


@login_required
def examenes_report(request):
    form = ExamenesFilterForm(request.GET) # Siempre inicializa el formulario con request.GET

    examenes = [] # Lista para almacenar los resultados del informe
    examenes_page_obj = None # Objeto de paginación
    report_executed = False # Bandera para saber si la consulta se ejecutó

    # Comprueba si el formulario ha sido enviado con datos (es decir, el usuario ha aplicado filtros)
    # y si esos datos son válidos
    if 'anio_academico' in request.GET or 'propuesta_ids' in request.GET:
        # Se ha enviado el formulario de filtrado, intentar ejecutar la consulta
        if form.is_valid():
            anio_academico_filter = form.cleaned_data.get('anio_academico')
            propuesta_ids_filter = form.cleaned_data.get('propuesta_ids')

            # Solo si se han proporcionado valores de filtro, ejecutar la consulta
            if anio_academico_filter and propuesta_ids_filter:
                report_executed = True

                # Convertir propuesta_ids_filter a una cadena separada por comas
                propuesta_ids_str = ','.join(map(str, [int(p_id) for p_id in propuesta_ids_filter]))

                with connection.cursor() as cursor:
                    sql_query = f"""
                            SELECT
                                    sa.nro_acta as acta,
                                    vme.anio_academico AS anio_academico,
                                    se2.nombre AS nombre_materia,
                                    se2.codigo as codigo,
                                    negocio_pers.f_cant_alumnos_acta(sa.id_acta) AS inscriptos,
                                    SUM(CASE WHEN ad.resultado = 'A' THEN 1 ELSE 0 END) AS aprobados,
                                    SUM(CASE WHEN ad.resultado = 'R' THEN 1 ELSE 0 END) AS reprobados,
                                    SUM(CASE WHEN ad.resultado = 'U' THEN 1 ELSE 0 END) AS ausentes
                            FROM
                                    negocio.sga_actas sa
                            LEFT JOIN
                                    negocio.sga_llamados_mesa slm ON slm.llamado_mesa = sa.llamado_mesa
                            LEFT JOIN
                                    negocio.sga_actas_detalle ad ON ad.id_acta = sa.id_acta
                            LEFT JOIN
                                    negocio.vw_mesas_examen vme ON vme.mesa_examen = slm.mesa_examen AND vme.llamado_mesa = sa.llamado_mesa
                            LEFT JOIN
                                    negocio.sga_elementos se2 ON se2.elemento = vme.mesa_examen_elemento
                            WHERE
                                    sa.origen = 'E'
                            AND vme.anio_academico = %s
                            AND sa.estado = 'C'
                            AND negocio_pers.get_propuesta_de_actividad(vme.mesa_examen_elemento) IN ({propuesta_ids_str})
                            GROUP BY
                                    sa.nro_acta,
                                    sa.id_acta,
                                    vme.anio_academico,
                                    se2.nombre,
                                    se2.codigo
                            ORDER BY
                                    sa.nro_acta, nombre_materia
                    """
                    cursor.execute(sql_query, [anio_academico_filter])

                    columns = [col[0] for col in cursor.description]
                    all_examenes_data = cursor.fetchall()

                examenes_list = []
                for row in all_examenes_data:
                    row_dict = dict(zip(columns, row))
                    examenes_list.append(row_dict)

                paginator = Paginator(examenes_list, 25)
                page_number = request.GET.get('page')

                try:
                    examenes_page_obj = paginator.page(page_number)
                except PageNotAnInteger:
                    examenes_page_obj = paginator.page(1)
                except EmptyPage:
                    examenes_page_obj = paginator.page(paginator.num_pages)

                # Si hay resultados de paginación, prepáralos para la visualización
                if examenes_page_obj:
                    for examen_dict in examenes_page_obj:
                        examenes.append({
                            'acta': examen_dict['acta'],
                            'anio_academico': examen_dict['anio_academico'],
                            'nombre_materia': examen_dict['nombre_materia'],
                            'codigo': examen_dict['codigo'],
                            'inscriptos': examen_dict['inscriptos'],
                            'aprobados': examen_dict['aprobados'],
                            'reprobados': examen_dict['reprobados'],
                            'ausentes': examen_dict['ausentes'],
                        })

    context = {
        'form': form,
        'examenes': examenes, # Ahora esta lista estará vacía si la consulta no se ejecuta
        'examenes_page_obj': examenes_page_obj,
        'report_title': 'Reporte de Exámenes',
        'report_executed': report_executed, # Pasa esta bandera al template
    }

    return render(request, 'account/reportes/examenes/examenes_report.html', context)


############################################################################
#CURSADAS REPORTES
#############################################################################


@login_required
def cursadas_report(request):
    form = CursadasFilterForm(request.GET) # Siempre inicializa el formulario con request.GET

    cursada = [] # Lista para almacenar los resultados del informe
    cursadas_page_obj = None # Objeto de paginación
    report_executed = False # Bandera para saber si la consulta se ejecutó

    # Comprueba si el formulario ha sido enviado con datos (es decir, el usuario ha aplicado filtros)
    # y si esos datos son válidos
    if 'anio' in request.GET or 'propuesta_ids' in request.GET:
        # Se ha enviado el formulario de filtrado, intentar ejecutar la consulta
        if form.is_valid():
            anio_filter = form.cleaned_data.get('anio')
            propuesta_ids_filter = form.cleaned_data.get('propuesta_ids')

            # Solo si se han proporcionado valores de filtro, ejecutar la consulta
            if anio_filter and propuesta_ids_filter:
                report_executed = True

                anio_filter = int(anio_filter)

                # Convertir propuesta_ids_filter a una cadena separada por comas
                propuesta_ids_str = ','.join(map(str, [int(p_id) for p_id in propuesta_ids_filter]))

                # --- LÍNEAS DE DEPURACIÓN CRÍTICAS AQUÍ ---
                #print(f"Valor de anio_filter ANTES de la ejecución de la consulta: {anio_filter}")
                #print(f"Tipo de anio_filter ANTES de la ejecución de la consulta: {type(anio_filter)}")
                # --- FIN LÍNEAS DE DEPURACIÓN ---

                with connection.cursor() as cursor:
                    sql_query = f"""
                            SELECT DISTINCT
                                    pr.codigo,
                                    sa.legajo,
                                    mp.apellido || ', ' || mp.nombres AS nombre_completo,
                                    mp.sexo as sexo,
                                    negocio_pers.f_edad_que_tenia_al(mp.fecha_nacimiento,MAKE_DATE(%s, EXTRACT(MONTH FROM NOW())::INT,
                                                                     EXTRACT(DAY FROM NOW())::INT)) as edad,
                                    negocio_pers.f_documento(mp.persona) AS dni,
                                    negocio_pers.get_mail(sa.persona) AS mail,
                                CASE
                                WHEN %s = EXTRACT(YEAR FROM CURRENT_DATE)  THEN negocio_pers.anio_que_cursa_actualmente(sa.propuesta, sa.alumno)
                                ELSE negocio_pers.anio_que_cursaba(sa.propuesta, sa.alumno, %s) -- Usamos $1 para el año que cursaba
                                END AS anio_que_cursa,
                                    t.ciudad as ciudad,
                                    t.cp as cp,
                                    t.provincia as provincia,
                                    t.pais as pais
                            FROM
                                    negocio.mdp_personas mp
                                    INNER JOIN negocio.sga_alumnos sa ON mp.persona = sa.persona
                                    INNER JOIN negocio.sga_propuestas pr ON sa.propuesta = pr.propuesta
                                    LEFT JOIN LATERAL negocio_pers.sp_domicilio_persona(mp.persona, 'PROC') as t ON TRUE
                            where
                                    pr.propuesta IN ({propuesta_ids_str})
                            AND EXISTS (
                                        SELECT 1
                                                FROM
                                                negocio.vw_comisiones c
                                                INNER JOIN negocio.sga_insc_cursada ic ON c.comision = ic.comision
                                        WHERE
                                                ic.alumno = sa.alumno
                                                AND c.anio_academico = %s
                                                AND ic.estado = 'A'
                                            )
                            order by nombre_completo
                    """
                    cursor.execute(sql_query, [anio_filter,anio_filter,anio_filter,anio_filter])

                    columns = [col[0] for col in cursor.description]
                    all_cursadas_data = cursor.fetchall()

                cursadas_list = []
                for row in all_cursadas_data:
                    row_dict = dict(zip(columns, row))
                    cursadas_list.append(row_dict)

                paginator = Paginator(cursadas_list, 25)
                page_number = request.GET.get('page')

                try:
                    cursadas_page_obj = paginator.page(page_number)
                except PageNotAnInteger:
                    cursadas_page_obj = paginator.page(1)
                except EmptyPage:
                    cursadas_page_obj = paginator.page(paginator.num_pages)

                # Si hay resultados de paginación, prepáralos para la visualización
                if cursadas_page_obj:
                    for cursada_dict in cursadas_page_obj:
                        cursada.append({
                            'codigo': cursada_dict['codigo'],
                            'legajo': cursada_dict['legajo'],
                            'nombre_completo': cursada_dict['nombre_completo'],
                            'sexo': cursada_dict['sexo'],
                            'edad': cursada_dict['edad'],
                            'dni': cursada_dict['dni'],
                            'mail': cursada_dict['mail'],
                            'anio_que_cursa': cursada_dict['anio_que_cursa'],
                            'ciudad': cursada_dict['ciudad'],
                            'cp': cursada_dict['cp'],
                            'provincia': cursada_dict['provincia'],
                            'pais': cursada_dict['pais'],
                        })

    context = {
        'form': form,
        'cursadas': cursada, # Ahora esta lista estará vacía si la consulta no se ejecuta
        'cursadas_page_obj': cursadas_page_obj,
        'report_title': 'Reporte de Cursadas',
        'report_executed': report_executed, # Pasa esta bandera al template
    }

    return render(request, 'account/reportes/cursadas/cursadas_report.html', context)



############################################################################
#PROMEDIOS REPORTES
#############################################################################

@login_required
def promedio_historico_report(request):
    report_executed = True
    promedios_by_year_and_generation = []
    paginator = None
    promedios_page_obj = None

    with connection.cursor() as cursor:
        sql_query = """
            SELECT ph.propuesta AS codigo_propuesta,
                   sp.nombre AS nombre_propuesta,
                   EXTRACT(YEAR FROM ph.fecha_generacion) AS anio,
                   ph.fecha_generacion AS generacion_raw,
                   ph.egresados_fecha_desde AS desde_raw,
                   ph.egresados_fecha_hasta AS hasta_raw,
                   ph.promedio_sin_aplazos AS sin_aplazos,
                   ph.promedio_general AS promedio,
                   ph.promedio_general_ponderado AS promedio_ponderado,
                   ph.desviacion_std AS desviacion,
                   ph.cant_egresados AS cantidad_egresados,
                   ph.cant_mat_total AS total_materias,
                   ph.cant_aprob AS materias_aprobadas
            FROM negocio.sga_propuestas sp,
                 negocio_pers.sga835_promedios_historicos ph
            WHERE sp.propuesta = ph.propuesta
            ORDER BY 3 DESC, 4 DESC, 1
        """
        cursor.execute(sql_query)

        columns = [col[0] for col in cursor.description]
        all_promedios_data = cursor.fetchall()

        promedios_list = []
        for row in all_promedios_data:
            row_dict = dict(zip(columns, row))

            # Conversión explícita a date object si es necesario
            # Si desde_raw/hasta_raw vienen como datetime, convertirlos a date
            if isinstance(row_dict['desde_raw'], datetime): # <--- Esta es la línea 406 (o similar)
                row_dict['desde'] = row_dict['desde_raw'].date()
            else:
                row_dict['desde'] = row_dict['desde_raw'] # Ya es date o None

            if isinstance(row_dict['hasta_raw'], datetime):
                row_dict['hasta'] = row_dict['hasta_raw'].date()
            else:
                row_dict['hasta'] = row_dict['hasta_raw'] # Ya es date o None

            # Mantener generacion como viene, probablemente es datetime
            row_dict['generacion'] = row_dict['generacion_raw']

            promedios_list.append(row_dict)

    # Group by 'anio' and then by 'generacion'
    promedios_list_sorted = sorted(promedios_list, key=itemgetter('anio', 'generacion'), reverse=True)

    for anio, year_group in groupby(promedios_list_sorted, key=itemgetter('anio')):
        generations_for_year = []
        for generacion, gen_group in groupby(list(year_group), key=itemgetter('generacion')):
            promedios_data_for_generation = []
            for row_dict in list(gen_group):
                promedios_data_for_generation.append({
                    'codigo': row_dict['codigo_propuesta'],
                    'carrera': row_dict['nombre_propuesta'],
                    'desde': row_dict['desde'], # Usa el campo ya convertido
                    'hasta': row_dict['hasta'], # Usa el campo ya convertido
                    'sin_aplazos': row_dict['sin_aplazos'],
                    'promedio': row_dict['promedio'],
                    'promedio_ponderado': row_dict['promedio_ponderado'],
                    'desviacion': row_dict['desviacion'],
                    'cantidad_egresados': row_dict['cantidad_egresados'],
                    'total_materias': row_dict['total_materias'],
                    'materias_aprobadas': row_dict['materias_aprobadas'],
                })
            generations_for_year.append({
                'generacion': generacion,
                'promedios_data': promedios_data_for_generation
            })
        promedios_by_year_and_generation.append({
            'anio': anio,
            'generations': generations_for_year
        })

    paginator = Paginator(promedios_by_year_and_generation, 5)
    page_number = request.GET.get('page')

    try:
        promedios_page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        promedios_page_obj = paginator.page(1)
    except EmptyPage:
        promedios_page_obj = paginator.page(paginator.num_pages)

    context = {
        'promedios_by_year_and_generation': promedios_page_obj,
        'report_title': 'Reporte de Promedios Históricos',
        'report_executed': report_executed,
    }

    return render(request, 'account/reportes/promedio_historico/promedio_historico_report.html', context)


############################################################################
#RESULTADO DE CURSADAS REPORTES
#############################################################################

@login_required
def resultado_cursadas_report(request):
    form = ResultadoCursaFilterForm(request.GET) # Siempre inicializa el formulario con request.GET

    resultado_cursada = [] # Lista para almacenar los resultados del informe
    resultado_cursadas_page_obj = None # Objeto de paginación
    report_executed = False # Bandera para saber si la consulta se ejecutó

    # Comprueba si el formulario ha sido enviado con datos (es decir, el usuario ha aplicado filtros)
    # y si esos datos son válidos
    if 'anio' in request.GET or 'propuesta_ids' in request.GET:
        # Se ha enviado el formulario de filtrado, intentar ejecutar la consulta
        if form.is_valid():
            anio_filter = form.cleaned_data.get('anio')
            propuesta_ids_filter = form.cleaned_data.get('propuesta_ids')

            # Solo si se han proporcionado valores de filtro, ejecutar la consulta
            if anio_filter and propuesta_ids_filter:
                report_executed = True

                # Convertir propuesta_ids_filter a una cadena separada por comas
                propuesta_ids_str = ','.join(map(str, [int(p_id) for p_id in propuesta_ids_filter]))

                with connection.cursor() as cursor:
                    sql_query = f"""
                    SET search_path TO negocio, negocio_pers;

                    SELECT
                            c.anio_academico AS anio,
                            c.comision_nombre AS nombre_comision,
                            c.comision AS comision,
                            c.elemento_nombre AS materia,
                            c.elemento_codigo AS codigo_materia,
                            CASE
                                WHEN ac.origen = 'R' THEN 'Acta_Regular'
                                WHEN ac.origen = 'P' THEN 'Acta_Promocion'
                            ELSE 'Acta_No_Valida'
                            END AS tipo_acta,
                        SUM(CASE WHEN negocio_pers.f_es_recursante(ad.alumno, c.anio_academico::integer, c.elemento, ic.fecha_inscripcion::date) = 'S' THEN 1 ELSE 0 END) AS recursantes_total,
                        SUM(CASE WHEN negocio_pers.f_es_recursante(ad.alumno, c.anio_academico::integer, c.elemento, ic.fecha_inscripcion::date) = 'N' THEN 1 ELSE 0 END) AS no_recursantes_total,
                        SUM(CASE WHEN negocio_pers.f_es_recursante(ad.alumno, c.anio_academico::integer, c.elemento, ic.fecha_inscripcion::date) = 'S' AND ad.resultado = 'A' THEN 1 ELSE 0 END) AS recursantes_aprobados,
                        SUM(CASE WHEN negocio_pers.f_es_recursante(ad.alumno, c.anio_academico::integer, c.elemento, ic.fecha_inscripcion::date) = 'S' AND ad.resultado != 'A' THEN 1 ELSE 0 END) AS recursantes_no_aprobados,
                        SUM(CASE WHEN negocio_pers.f_es_recursante(ad.alumno, c.anio_academico::integer, c.elemento, ic.fecha_inscripcion::date) = 'N' AND ad.resultado = 'A' THEN 1 ELSE 0 END) AS no_recursantes_aprobados,
                        SUM(CASE WHEN negocio_pers.f_es_recursante(ad.alumno, c.anio_academico::integer, c.elemento, ic.fecha_inscripcion::date) = 'N' AND ad.resultado != 'A' THEN 1 ELSE 0 END) AS no_recursantes_no_aprobados,
                        negocio_pers.get_docentes_de_una_comision(c.comision) as docentes
                    FROM
                        negocio.vw_comisiones c
                    INNER JOIN
                        negocio.sga_actas ac ON c.comision = ac.comision
                    INNER JOIN
                        negocio.sga_actas_detalle ad ON ac.id_acta = ad.id_acta
                    INNER JOIN
                        negocio.sga_insc_cursada ic ON c.comision = ic.comision AND ad.alumno = ic.alumno
                    INNER JOIN
                        negocio.sga_alumnos a ON ad.alumno = a.alumno -- Aunque 'a.legajo' no se usa en el SELECT final, mantenemos el JOIN por si hay alguna dependencia implícita.
                    WHERE
                        c.anio_academico = %s
                        AND negocio_pers.get_propuesta_de_actividad(c.elemento) IN ({propuesta_ids_str})
                        AND ac.estado = 'C'
                    GROUP BY
                        c.anio_academico,
                        c.comision_nombre,
                        c.comision,
                        c.elemento_nombre,
                        c.elemento_codigo,
                        CASE
                            WHEN ac.origen = 'R' THEN 'Acta_Regular'
                            WHEN ac.origen = 'P' THEN 'Acta_Promocion'
                            ELSE 'Acta_No_Valida'
                        END
                    ORDER BY
                        c.anio_academico, c.comision_nombre, tipo_acta
                    """
                    cursor.execute(sql_query, [anio_filter])

                    columns = [col[0] for col in cursor.description]
                    all_resultado_cursadas_data = cursor.fetchall()

                resultado_cursadas_list = []
                for row in all_resultado_cursadas_data:
                    row_dict = dict(zip(columns, row))
                    resultado_cursadas_list.append(row_dict)

                paginator = Paginator(resultado_cursadas_list, 25)
                page_number = request.GET.get('page')

                try:
                    resultado_cursadas_page_obj = paginator.page(page_number)
                except PageNotAnInteger:
                    resultado_cursadas_page_obj = paginator.page(1)
                except EmptyPage:
                    resultado_cursadas_page_obj = paginator.page(paginator.num_pages)

                # Si hay resultados de paginación, prepáralos para la visualización
                if resultado_cursadas_page_obj:
                    for resultado_cursada_dict in resultado_cursadas_page_obj:
                        resultado_cursada.append({
                            'anio': resultado_cursada_dict['anio'],
                            'nombre_comision': resultado_cursada_dict['nombre_comision'],
                            'comision': resultado_cursada_dict['comision'],
                            'materia': resultado_cursada_dict['materia'],
                            'codigo_materia': resultado_cursada_dict['codigo_materia'],
                            'tipo_acta': resultado_cursada_dict['tipo_acta'],
                            'recursantes_total': resultado_cursada_dict['recursantes_total'],
                            'no_recursantes_total': resultado_cursada_dict['no_recursantes_total'],
                            'recursantes_aprobados': resultado_cursada_dict['recursantes_aprobados'],
                            'recursantes_no_aprobados': resultado_cursada_dict['recursantes_no_aprobados'],
                            'no_recursantes_aprobados': resultado_cursada_dict['no_recursantes_aprobados'],
                            'no_recursantes_no_aprobados': resultado_cursada_dict['no_recursantes_no_aprobados'],
                            'docentes': resultado_cursada_dict['docentes'],
                        })
    context = {
        'form': form,
        'resultado_cursada': resultado_cursada, # Ahora esta lista estará vacía si la consulta no se ejecuta
        'resultado_cursadas_page_obj': resultado_cursadas_page_obj,
        'report_title': 'Reporte de Resultado de Cursadas',
        'report_executed': report_executed, # Pasa esta bandera al template
    }

    return render(request, 'account/reportes/resultado_cursada/resultado_cursada_report.html', context)

#########################################################################
#################### ESTADISTICAS #######################################
############################################################################

#########################################################################
# INGRESANTES POR CARRERA Y PLAN
############################################################################

@login_required
def ingresantes_por_carrera_view(request):
    form = IngresantesFilterForm(request.GET)

    # Inicializa todas las variables relacionadas con los datos a vacío/None
    labels = []
    data = []
    background_colors = []
    processed_columns = []
    results = []
    anio_filter = None
    # Esta bandera indicará si se generó un reporte (después de aplicar filtros)
    report_generated = False

    # Solo intenta procesar datos si hay parámetros GET en la solicitud
    # Esto implica que el formulario fue enviado
    if request.GET:
        # Si el formulario es válido, significa que se enviaron filtros
        if form.is_valid():
            report_generated = True # La bandera se establece en True

            anio_filter = form.cleaned_data.get('anio')
            propuesta_ids_filter = form.cleaned_data.get('propuesta_ids')

            if anio_filter:
                try:
                    anio_filter = int(anio_filter)
                except (ValueError, TypeError):
                    anio_filter = None

            propuesta_ids_filter = [int(p_id) for p_id in propuesta_ids_filter if p_id]

            sql_where_clauses = ["pv.estado = 'V'"]
            sql_params = []

            if anio_filter:
                sql_where_clauses.append("negocio_pers.anioingreso(a.alumno) = %s")
                sql_params.append(anio_filter)

            if propuesta_ids_filter:
                placeholders = ','.join(['%s'] * len(propuesta_ids_filter))
                sql_where_clauses.append(f"sp.propuesta IN ({placeholders})")
                sql_params.extend(propuesta_ids_filter)

            where_clause_str = " AND ".join(sql_where_clauses)
            if where_clause_str:
                where_clause_str = "WHERE " + where_clause_str

            try:
                with connection.cursor() as cursor:
                    sql_query = f"""
                        SELECT
                            p.nombre AS plan_estudios,
                            COUNT(a.alumno) AS cantidad_alumnos,
                            negocio_pers.anioingreso(a.alumno) AS anio_ingreso,
                            sp.nombre AS carrera
                        FROM
                            negocio.sga_alumnos a
                        INNER JOIN
                            negocio.sga_propuestas sp ON a.propuesta = sp.propuesta
                        INNER JOIN
                            negocio.sga_planes_versiones pv ON a.plan_version = pv.plan_version
                        INNER JOIN
                            negocio.sga_planes p ON pv.plan = p.plan AND sp.propuesta = p.propuesta
                        {where_clause_str}
                        GROUP BY
                            p.nombre,
                            negocio_pers.anioingreso(a.alumno),
                            sp.nombre
                        ORDER BY
                            p.nombre,
                            negocio_pers.anioingreso(a.alumno) ASC;
                    """
                    cursor.execute(sql_query, sql_params)
                    columns = [col[0] for col in cursor.description]
                    results = cursor.fetchall()

                # Procesa los resultados para Chart.js solo si hay resultados
                if results:
                    colors = [
                        'rgba(255, 99, 132, 0.6)', 'rgba(54, 162, 235, 0.6)', 'rgba(255, 206, 86, 0.6)',
                        'rgba(75, 192, 192, 0.6)', 'rgba(153, 102, 255, 0.6)', 'rgba(255, 159, 64, 0.6)',
                        'rgba(199, 199, 199, 0.6)', 'rgba(83, 102, 255, 0.6)', 'rgba(120, 180, 240, 0.6)'
                    ]
                    color_index = 0

                    for row in results:
                        plan_estudios, cantidad_alumnos, anio_ingreso_resultado, carrera = row
                        labels.append(f"{carrera} ({plan_estudios})")
                        data.append(cantidad_alumnos)
                        background_colors.append(colors[color_index % len(colors)])
                        color_index += 1

                processed_columns = [col_name.replace('_', ' ').capitalize() for col_name in columns]

            except Exception as e:
                print(f"Error en ingresantes_por_carrera_view: {e}")
                # Si ocurre un error, asegúrate de que los datos para el reporte estén vacíos
                labels = []
                data = []
                background_colors = []
                processed_columns = []
                results = []
                report_generated = False # Si hay un error, no se considera que se generó un reporte válido
                form.add_error(None, f"No se pudo cargar el reporte. Error: {e}")
        else:
            # Si el formulario NO es válido (ej. filtros mal formados),
            # no se generará un reporte y se mostrarán los errores del formulario.
            # report_generated ya es False por defecto.
            pass # No hacemos nada aquí, ya que el formulario maneja sus propios errores
    else:
        # Si no hay request.GET (primera carga), no se genera ningún reporte
        # report_generated ya es False por defecto.
        pass


    context = {
        'form': form,
        # Solo muestra el año filtrado si se generó un reporte, de lo contrario, "Todos"
        'anio_ingreso_filtrado': anio_filter if anio_filter else 'Todos',
        'labels': labels,
        'data': data,
        'background_colors': background_colors,
        'report_data': results,
        'columns': processed_columns,
        'report_generated': report_generated, # Pasa esta bandera al template
    }

    return render(request, 'account/estadisticas/ingresantes/ingresantes_report.html', context)



#########################################################################
# DOCENTES POR CARRERA O DEPARTAMENTO
############################################################################

@login_required
def docentes_x_carrera_dpto_view(request):
    form = DocentesFilterForm(request.GET)

    labels = []
    data_unicos = []
    data_totales = []
    background_colors_unicos = []
    background_colors_totales = []
    processed_columns = []
    results = []
    anio_filter = None
    report_generated = False

    # Listas para almacenar las cadenas SQL de cada rama y sus parámetros individuales
    union_query_strings = []
    union_query_params_individual = [] # Para acumular los parámetros de cada rama

    if request.GET:
        if form.is_valid():
            report_generated = True

            anio_filter = form.cleaned_data.get('anio')
            propuesta_ids_filter = form.cleaned_data.get('propuesta_ids') # .cleaned_data es correcto
            dptos_ids_filter = form.cleaned_data.get('dptos_ids')

            if propuesta_ids_filter:
                propuesta_ids_filter = [int(p_id) for p_id in propuesta_ids_filter if p_id]
            else:
                propuesta_ids_filter = []

            if dptos_ids_filter:
                dptos_ids_filter = [int(d_id) for d_id in dptos_ids_filter if d_id]
            else:
                dptos_ids_filter = []

            try:
                with connection.cursor() as cursor:
                    # --- Construcción Dinámica de las Ramas SQL y sus Parámetros ---

                    # Lógica para la Rama de Propuestas
                    # Incluir esta rama si:
                    # 1. Hay filtros de propuesta seleccionados, O
                    # 2. NO hay filtros de propuesta Y NO hay filtros de departamento (es decir, mostrar ambas por defecto)
                    if propuesta_ids_filter or (not propuesta_ids_filter and not dptos_ids_filter):
                        propuesta_branch_where_clauses = ["c.estado = 'A'"]
                        propuesta_branch_params = []

                        if anio_filter:
                            propuesta_branch_where_clauses.append("c.anio_academico = %s")
                            propuesta_branch_params.append(anio_filter)

                        if propuesta_ids_filter:
                            propuesta_branch_where_clauses.append(f"pro.propuesta IN ({','.join(['%s'] * len(propuesta_ids_filter))})")
                            propuesta_branch_params.extend(propuesta_ids_filter)

                        union_query_strings.append(f"""
                            SELECT
                                'PROPUESTA' AS tipo_filtro,
                                pro.nombre AS carrera,
                                pro.codigo AS codigo_carrera,
                                COUNT(DISTINCT dc.docente) AS cantidad_docentes_unicos_por_carrera,
                                negocio_pers.get_cant_docentesduplicados(c.anio_academico::integer, pro.propuesta, NULL) AS cantidad_docentes_por_carrera
                            FROM
                                negocio.vw_comisiones AS c
                            INNER JOIN negocio.sga_docentes_comision AS dc ON c.comision = dc.comision
                            INNER JOIN negocio.sga_docentes AS d ON dc.docente = d.docente
                            INNER JOIN negocio.mdp_personas AS p ON d.persona = p.persona
                            INNER JOIN negocio.sga_elementos AS e ON c.elemento = e.elemento
                            INNER JOIN negocio.sga_elementos_revision AS er ON e.elemento = er.elemento
                            INNER JOIN negocio.sga_elementos_plan AS ep ON er.elemento_revision = ep.elemento_revision
                            INNER JOIN negocio.sga_planes_versiones AS pv ON ep.plan_version = pv.plan_version
                            INNER JOIN negocio.sga_planes AS pr ON pv.plan = pr.plan
                            INNER JOIN negocio.sga_propuestas AS pro ON pr.propuesta = pro.propuesta
                            INNER JOIN negocio.sga_elementos_ra AS era ON e.elemento = era.elemento
                            INNER JOIN negocio.sga_responsables_academicas AS ra ON ra.responsable_academica = era.responsable_academica AND ra.responsable_academica_tipo = 2
                            WHERE
                                {' AND '.join(propuesta_branch_where_clauses)}
                            GROUP BY
                                tipo_filtro,
                                carrera,
                                codigo_carrera,
                                c.anio_academico,
                                pro.propuesta
                        """)
                        union_query_params_individual.extend(propuesta_branch_params)


                    # Lógica para la Rama de Departamentos (Responsables Académicos)
                    # Incluir esta rama si:
                    # 1. Hay filtros de departamento seleccionados, O
                    # 2. NO hay filtros de propuesta Y NO hay filtros de departamento (es decir, mostrar ambas por defecto)
                    if dptos_ids_filter or (not propuesta_ids_filter and not dptos_ids_filter):
                        dpto_branch_where_clauses = ["c.estado = 'A'"]
                        dpto_branch_params = []

                        if anio_filter:
                            dpto_branch_where_clauses.append("c.anio_academico = %s")
                            dpto_branch_params.append(anio_filter)

                        if dptos_ids_filter:
                            dpto_branch_where_clauses.append(f"ra.responsable_academica IN ({','.join(['%s'] * len(dptos_ids_filter))})")
                            dpto_branch_params.extend(dptos_ids_filter)

                        union_query_strings.append(f"""
                            SELECT
                                'DEPARTAMENTO' AS tipo_filtro,
                                ra.nombre AS carrera,
                                ra.codigo AS codigo_carrera,
                                COUNT(DISTINCT dc.docente) AS cantidad_docentes_unicos_por_carrera,
                                negocio_pers.get_cant_docentesduplicados(c.anio_academico::integer, NULL, ra.responsable_academica) AS cantidad_docentes_por_carrera
                            FROM
                                negocio.vw_comisiones AS c
                            INNER JOIN negocio.sga_docentes_comision AS dc ON c.comision = dc.comision
                            INNER JOIN negocio.sga_docentes AS d ON dc.docente = d.docente
                            INNER JOIN negocio.mdp_personas AS p ON d.persona = p.persona
                            INNER JOIN negocio.sga_elementos AS e ON c.elemento = e.elemento
                            INNER JOIN negocio.sga_elementos_revision AS er ON e.elemento = er.elemento
                            INNER JOIN negocio.sga_elementos_plan AS ep ON er.elemento_revision = ep.elemento_revision
                            INNER JOIN negocio.sga_planes_versiones AS pv ON ep.plan_version = pv.plan_version
                            INNER JOIN negocio.sga_planes AS pr ON pv.plan = pr.plan
                            INNER JOIN negocio.sga_propuestas AS pro ON pr.propuesta = pro.propuesta
                            INNER JOIN negocio.sga_elementos_ra AS era ON e.elemento = era.elemento
                            INNER JOIN negocio.sga_responsables_academicas AS ra ON ra.responsable_academica = era.responsable_academica AND ra.responsable_academica_tipo = 2
                            WHERE
                                {' AND '.join(dpto_branch_where_clauses)}
                            GROUP BY
                                tipo_filtro,
                                carrera,
                                codigo_carrera,
                                c.anio_academico,
                                ra.responsable_academica
                        """)
                        union_query_params_individual.extend(dpto_branch_params)


                    # --- Ejecución de la Consulta ---
                    if not union_query_strings:
                        # Si ninguna rama fue seleccionada (ej. si los filtros específicos no están y no se desea el default)
                        report_generated = False
                        # Opcional: form.add_error(None, "Por favor, selecciona al menos un filtro de Carrera o Departamento.")
                    else:
                        final_sql_query = " UNION ALL ".join(union_query_strings) + " ORDER BY tipo_filtro, carrera, codigo_carrera;"

                        # Los parámetros finales son la concatenación de todos los parámetros individuales
                        final_sql_params = union_query_params_individual

                        cursor.execute(final_sql_query, final_sql_params)
                        columns = [col[0] for col in cursor.description]
                        results = cursor.fetchall()

                        # Procesa los resultados para Chart.js solo si hay resultados
                        if results:
                            color_unicos = 'rgba(54, 162, 235, 0.6)'
                            color_totales = 'rgba(255, 99, 132, 0.6)'

                            for row in results:
                                tipo_filtro, carrera_nombre, codigo_carrera_val, cant_unicos, cant_totales = row
                                labels.append(f"{carrera_nombre} ({tipo_filtro})")
                                data_unicos.append(cant_unicos)
                                data_totales.append(cant_totales)
                                background_colors_unicos.append(color_unicos)
                                background_colors_totales.append(color_totales)

                    processed_columns = [col_name.replace('_', ' ').capitalize() for col_name in columns]

            except psycopg2.Error as e:
                print(f"Error en la consulta SQL de docentes_x_carrera_dpto_view: {e}")
                labels = []
                data_unicos = []
                data_totales = []
                background_colors_unicos = []
                background_colors_totales = []
                processed_columns = []
                results = []
                report_generated = False
                form.add_error(None, f"No se pudo cargar el reporte debido a un error en la base de datos: {e}")
            except Exception as e:
                print(f"Error inesperado en docentes_x_carrera_dpto_view: {e}")
                labels = []
                data_unicos = []
                data_totales = []
                background_colors_unicos = []
                background_colors_totales = []
                processed_columns = []
                results = []
                report_generated = False
                form.add_error(None, f"No se pudo cargar el reporte. Error: {e}")
        else:
            pass # El formulario no es válido, los errores ya están en form.errors
    else:
        pass # No hay parámetros GET, es la primera carga de la página

    context = {
        'form': form,
        'anio_ingreso_filtrado': anio_filter if anio_filter else 'Todos',
        'labels': labels,
        'data_unicos': data_unicos,
        'data_totales': data_totales,
        'background_colors_unicos': background_colors_unicos,
        'background_colors_totales': background_colors_totales,
        'report_data': results,
        'columns': processed_columns,
        'report_generated': report_generated,
    }

    return render(request, 'account/estadisticas/docentes/docentes_report.html', context)

#########################################################################
# RETENIDOS POR CARRERA ALUMNOS DE PRIMER AÑO
############################################################################

@login_required
def retenidos_por_carrera_view(request):
    form = IngresantesFilterForm(request.GET)

    labels = []
    data_ingresantes = []
    data_retenidos = []
    background_colors_ingresantes = []
    background_colors_retenidos = []
    processed_columns = []
    results = []
    anio_filter = None
    report_generated = False

    if request.GET:
        if form.is_valid():
            report_generated = True

            anio_filter = form.cleaned_data.get('anio')
            propuesta_ids_filter = form.cleaned_data.get('propuesta_ids') # Corregido a .cleaned_data

            # Asegúrate de que anio_filter sea un entero para la suma
            if anio_filter:
                try:
                    anio_filter = int(anio_filter)
                    anio_siguiente = anio_filter + 1 # <--- ¡CALCULA EL AÑO SIGUIENTE AQUÍ!
                except (ValueError, TypeError):
                    anio_filter = None
                    anio_siguiente = None # Si el año no es válido, tampoco lo será el siguiente
            else:
                anio_siguiente = None # Si no hay anio_filter, no hay anio_siguiente

            propuesta_ids_filter = [int(p_id) for p_id in propuesta_ids_filter if p_id]

            sql_where_clauses = [] # Comienza vacía, no con ""
            sql_params = []

            if anio_filter:
                sql_where_clauses.append("negocio_pers.anioingreso(sa.alumno) = %s")
                sql_params.append(anio_filter)

            if propuesta_ids_filter:
                placeholders = ','.join(['%s'] * len(propuesta_ids_filter))
                sql_where_clauses.append(f"pr.propuesta IN ({placeholders})")
                sql_params.extend(propuesta_ids_filter)

            where_clause_str = ""
            if sql_where_clauses: # Solo agrega WHERE si hay cláusulas
                where_clause_str = "WHERE " + " AND ".join(sql_where_clauses)

            if anio_filter is None:
                report_generated = False
                form.add_error(None, "Por favor, selecciona un año válido para generar el reporte.")
                # Saltamos el resto de la lógica de la base de datos
                context = {
                    'form': form,
                    'anio_ingreso_filtrado': 'Todos',
                    'labels': [],
                    'data_unicos': [],
                    'data_totales': [],
                    'background_colors_unicos': [],
                    'background_colors_totales': [],
                    'report_data': [],
                    'columns': [],
                    'report_generated': False,
                }
                return render(request, 'account/estadisticas/retenidos/retenidos_report.html', context)


            try:
                with connection.cursor() as cursor:
                    sql_query = f"""
                        SELECT
                            pr.nombre AS nombre_carrera,
                            pr.codigo AS codigo_carrera,
                            COUNT(sa.legajo) AS total_ingresantes,
                            COUNT(sa.legajo) FILTER (
                                WHERE
                                    EXISTS (
                                        SELECT 1
                                        FROM
                                            negocio.vw_comisiones c
                                        INNER JOIN negocio.sga_insc_cursada ic ON c.comision = ic.comision
                                        WHERE
                                            ic.alumno = sa.alumno
                                            AND c.anio_academico = %s
                                            AND ic.estado = 'A'
                                    )
                            ) AS alumnos_retenidos
                        FROM
                            negocio.mdp_personas mp
                        INNER JOIN negocio.sga_alumnos sa ON mp.persona = sa.persona
                        INNER JOIN negocio.sga_propuestas pr ON sa.propuesta = pr.propuesta
                        {where_clause_str}
                        GROUP BY
                            pr.nombre,
                            pr.codigo
                        ORDER BY
                            pr.nombre;
                    """
                    # Asegúrate de que anio_siguiente esté en sql_params
                    # Lo ideal es que vaya al principio si es el primer %s
                    # de la subconsulta del FILTER.
                    # El orden de los parámetros es CRUCIAL.

                    # Para la subconsulta, anio_academico es el primer %s en el SELECT.
                    # Luego vienen los %s de anioingreso y los %s de pr.propuesta IN (...).
                    # Por lo tanto, anio_siguiente debe ser el PRIMER parámetro en la lista.

                    # Vamos a crear la lista de parámetros final de forma explícita:
                    final_sql_params = [anio_siguiente] # El primer parámetro para el c.anio_academico del FILTER
                    final_sql_params.extend(sql_params) # Luego, los parámetros del WHERE principal (anio_filter, propuesta_ids)

                    cursor.execute(sql_query, final_sql_params)
                    columns = [col[0] for col in cursor.description]
                    results = cursor.fetchall()

                # Procesa los resultados para Chart.js
                if results: # Se movió el 'if results' para englobar el for loop
                    color_ingresantes = 'rgba(54, 162, 235, 0.6)'
                    color_retenidos = 'rgba(255, 99, 132, 0.6)'

                    for row in results:
                        # Asegúrate del orden de las columnas en el SELECT
                        nombre_carrera, codigo_carrera, total_ingresantes, alumnos_retenidos = row
                        labels.append(f"{nombre_carrera} ({codigo_carrera})")
                        data_ingresantes.append(total_ingresantes)
                        data_retenidos.append(alumnos_retenidos)
                        background_colors_ingresantes.append(color_ingresantes)
                        background_colors_retenidos.append(color_retenidos)

                processed_columns = [col_name.replace('_', ' ').capitalize() for col_name in columns]

            except psycopg2.Error as e: # Es bueno capturar errores específicos de la DB
                print(f"Error de base de datos en retenidos_por_carrera_view: {e}")
                labels = []
                data_ingresantes = []
                data_retenidos = []
                background_colors_ingresantes = []
                background_colors_retenidos = []
                processed_columns = []
                results = []
                report_generated = False
                form.add_error(None, f"No se pudo cargar el reporte debido a un error en la base de datos: {e}")
            except Exception as e:
                print(f"Error inesperado en retenidos_por_carrera_view: {e}")
                labels = []
                data_ingresantes = []
                data_retenidos = []
                background_colors_ingresantes = []
                background_colors_retenidos = []
                processed_columns = []
                results = []
                report_generated = False
                form.add_error(None, f"No se pudo cargar el reporte. Error: {e}")
        else:
            pass
    else:
        pass


    context = {
        'form': form,
        'anio_ingreso_filtrado': anio_filter if anio_filter else 'Todos',
        'labels': labels,
        'data_ingresantes': data_ingresantes,
        'data_retenidos': data_retenidos,
        'background_colors_ingresantes': background_colors_ingresantes,
        'background_colors_retenidos': background_colors_retenidos,
        'report_data': results,
        'columns': processed_columns,
        'report_generated': report_generated,
    }

    return render(request, 'account/estadisticas/retenidos/retenidos_report.html',context)



############################################################################
#EXAMENES EN EXCEL
#############################################################################

@login_required
def export_examenes_excel(request):
    # Inicializa el formulario con los parámetros GET recibidos
    form = ExamenesFilterForm(request.GET)

    # Valores predeterminados (por si acaso no se aplican filtros o son inválidos)
    anio_academico_filter = None
    propuesta_ids_filter = []

    # Validar los datos del formulario (los mismos filtros que en la vista del reporte)
    if form.is_valid():
        anio_academico_filter = form.cleaned_data.get('anio_academico')
        propuesta_ids_filter = form.cleaned_data.get('propuesta_ids')

    # Construir la consulta SQL dinámicamente basada en los filtros
    sql_where_clauses = ["sa.origen = 'E'", "sa.estado = 'C'"]
    sql_params = []

    if anio_academico_filter:
        sql_where_clauses.append("vme.anio_academico = %s")
        sql_params.append(anio_academico_filter)

    if propuesta_ids_filter:
        # Crea una cadena de marcadores de posición (%s) para la cláusula IN
        placeholders = ','.join(['%s'] * len(propuesta_ids_filter))
        sql_where_clauses.append(f"negocio_pers.get_propuesta_de_actividad(vme.mesa_examen_elemento) IN ({placeholders})")
        # Asegúrate de que los IDs sean enteros si tu base de datos los espera así
        sql_params.extend([int(p_id) for p_id in propuesta_ids_filter])


    # Unir las cláusulas WHERE
    where_clause_str = " AND ".join(sql_where_clauses)
    if where_clause_str: # Solo añade WHERE si hay cláusulas
        where_clause_str = "WHERE " + where_clause_str

    with connection.cursor() as cursor:
        sql_query = f"""
                SELECT
                        sa.nro_acta as acta,
                        vme.anio_academico AS anio_academico,
                        se2.nombre AS nombre_materia,
                        se2.codigo as codigo,
                        negocio_pers.f_cant_alumnos_acta(sa.id_acta) AS inscriptos,
                        SUM(CASE WHEN ad.resultado = 'A' THEN 1 ELSE 0 END) AS aprobados,
                        SUM(CASE WHEN ad.resultado = 'R' THEN 1 ELSE 0 END) AS reprobados,
                        SUM(CASE WHEN ad.resultado = 'U' THEN 1 ELSE 0 END) AS ausentes
                FROM
                        negocio.sga_actas sa
                LEFT JOIN
                        negocio.sga_llamados_mesa slm ON slm.llamado_mesa = sa.llamado_mesa
                LEFT JOIN
                        negocio.sga_actas_detalle ad ON ad.id_acta = sa.id_acta
                LEFT JOIN
                        negocio.vw_mesas_examen vme ON vme.mesa_examen = slm.mesa_examen AND vme.llamado_mesa = sa.llamado_mesa
                LEFT JOIN
                        negocio.sga_elementos se2 ON se2.elemento = vme.mesa_examen_elemento
                {where_clause_str}
                GROUP BY
                        sa.nro_acta,
                        sa.id_acta,
                        vme.anio_academico,
                        se2.nombre,
                        se2.codigo
                ORDER BY
                        sa.nro_acta, nombre_materia
        """
        cursor.execute(sql_query, sql_params) # Pasa los parámetros a la ejecución de la consulta
        columns = [col[0] for col in cursor.description]
        examenes_data = cursor.fetchall() # Cambiado de egresados_data a examenes_data

    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Examenes"

    # Escribir los encabezados (columnas de la consulta SQL)
    ws.append(columns)

    # Escribir los datos
    for row in examenes_data:
        processed_row = []
        for cell in row:
            if isinstance(cell, (date, datetime)):
                processed_row.append(cell.strftime('%Y-%m-%d'))
            else:
                processed_row.append(cell)
        ws.append(processed_row)

    # Configurar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_examenes.xlsx"'
    wb.save(response) # Guarda el libro de trabajo en la respuesta HTTP
    return response # No se retorna una tupla, solo el objeto HttpResponse


############################################################################
#EGRESADOS EN EXCEL
#############################################################################

@login_required
def export_egresados_excel(request):
    # Inicializa el formulario con los parámetros GET recibidos
    form = ExamenesFilterForm(request.GET)

    # Valores predeterminados (por si acaso no se aplican filtros o son inválidos)
    anio_filter = None
    propuesta_ids_filter = []

    # Validar los datos del formulario (los mismos filtros que en la vista del reporte)
    if form.is_valid():
        anio_filter = form.cleaned_data.get('anio')
        propuesta_ids_filter = form.cleaned_data.get('propuesta_ids')

    # Construir la consulta SQL dinámicamente basada en los filtros
    sql_where_clauses = ["a.persona = vp.persona", "a.alumno = sco.alumno", "vp.persona = sco.persona", "vp.nacionalidad = n.nacionalidad"]
    sql_params = []

    if anio_filter:
        sql_where_clauses.append("EXTRACT(YEAR FROM sco.fecha_egreso) = %s")
        sql_params.append(anio_filter)

    if propuesta_ids_filter:
        # Crea una cadena de marcadores de posición (%s) para la cláusula IN
        placeholders = ','.join(['%s'] * len(propuesta_ids_filter))
        sql_where_clauses.append(f"a.propuesta in ({placeholders})")
        # Asegúrate de que los IDs sean enteros si tu base de datos los espera así
        sql_params.extend([int(p_id) for p_id in propuesta_ids_filter])

    # Unir las cláusulas WHERE
    where_clause_str = " AND ".join(sql_where_clauses)
    if where_clause_str: # Solo añade WHERE si hay cláusulas
        where_clause_str = "WHERE " + where_clause_str

    with connection.cursor() as cursor:
        sql_query = f"""
                        select vp.apellido_nombres as nombre, n.descripcion as nacionalidad, negocio_pers.anioingreso(a.alumno) as anioingreso,
                               negocio_pers.get_mail(a.persona) as mail,vp.tipo_nro_documento as dni,  vp.sexo as sexo, a.legajo as legajo, a.propuesta_nombre as carrera,
                               a.plan_codigo as codigo, sco.fecha_egreso as egreso, sco.promedio as promedio, sco.promedio_sin_aplazos as sin_aplazos
                        from negocio.vw_alumnos a,
                                negocio.sga_certificados_otorg sco,
                                negocio.mdp_nacionalidades n ,
                                negocio.vw_personas vp
                        {where_clause_str}
                        ORDER BY
                                vp.apellido_nombres
        """
        cursor.execute(sql_query, sql_params) # Pasa los parámetros a la ejecución de la consulta
        columns = [col[0] for col in cursor.description]
        egresados_data = cursor.fetchall()

    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Egresados"

    # Escribir los encabezados (columnas de la consulta SQL)
    ws.append(columns)

    # Escribir los datos
    for row in egresados_data:
        processed_row = []
        for cell in row:
            if isinstance(cell, (date, datetime)):
                processed_row.append(cell.strftime('%Y-%m-%d'))
            else:
                processed_row.append(cell)
        ws.append(processed_row)

    # Configurar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_examenes.xlsx"'
    wb.save(response) # Guarda el libro de trabajo en la respuesta HTTP
    return response # No se retorna una tupla, solo el objeto HttpResponse


############################################################################
#CURSADAS EN EXCEL
#############################################################################

@login_required
def export_cursadas_excel(request):
    # Inicializa el formulario con los parámetros GET recibidos
    form = CursadasFilterForm(request.GET)

    # Valores predeterminados (por si acaso no se aplican filtros o son inválidos)
    anio_filter = None
    propuesta_ids_filter = []

    # Validar los datos del formulario (los mismos filtros que en la vista del reporte)
    if form.is_valid():
        anio_filter = form.cleaned_data.get('anio')
        propuesta_ids_filter = form.cleaned_data.get('propuesta_ids')

    # Construir la consulta SQL dinámicamente basada en los filtros
    sql_where_clauses = [] # Inicializa como vacío
    sql_params = []

    # --- Manejo de parámetros para la cláusula CASE en el SELECT ---
    case_params = []
    if anio_filter:
        case_params.append(anio_filter) # Primer %s para el WHEN
        case_params.append(anio_filter) # Segundo %s para la función anio_que_cursaba
        case_params.append(anio_filter) # Tercero %s para la función anio_que_cursaba
    # Si anio_filter no está presente, considera qué valor predeterminado o manejo quieres para el CASE.
    # Si 'anio' es un campo requerido en tu CursadasFilterForm, siempre estará aquí.

    if anio_filter:
        sql_where_clauses.append("EXISTS (SELECT 1 FROM negocio.vw_comisiones c INNER JOIN negocio.sga_insc_cursada ic ON c.comision = ic.comision WHERE ic.alumno = sa.alumno AND c.anio_academico = %s AND ic.estado = 'A')")
        sql_params.append(anio_filter)

    if propuesta_ids_filter:
        # Crea una cadena de marcadores de posición (%s) para la cláusula IN
        placeholders = ','.join(['%s'] * len(propuesta_ids_filter))
        sql_where_clauses.append(f"pr.propuesta in ({placeholders})")
        # Asegúrate de que los IDs sean enteros si tu base de datos los espera así
        sql_params.extend([int(p_id) for p_id in propuesta_ids_filter])

    # Unir las cláusulas WHERE
    where_clause_str = ""
    if sql_where_clauses: # Solo añade WHERE si hay cláusulas
        where_clause_str = "WHERE " + " AND ".join(sql_where_clauses)

    with connection.cursor() as cursor:
        sql_query = f"""
                        SELECT DISTINCT
                                    pr.codigo,
                                    sa.legajo,
                                    mp.apellido || ', ' || mp.nombres AS nombre_completo,
                                    mp.sexo as sexo,
                                    negocio_pers.f_edad_que_tenia_al(mp.fecha_nacimiento,MAKE_DATE(%s, EXTRACT(MONTH FROM NOW())::INT,
                                                                     EXTRACT(DAY FROM NOW())::INT)) as edad,
                                    negocio_pers.f_documento(mp.persona) AS dni,
                                    negocio_pers.get_mail(sa.persona) AS mail,
                                CASE
                                WHEN %s = EXTRACT(YEAR FROM CURRENT_DATE)  THEN negocio_pers.anio_que_cursa_actualmente(sa.propuesta, sa.alumno)
                                ELSE negocio_pers.anio_que_cursaba(sa.propuesta, sa.alumno, %s)
                                END AS anio_que_cursa,
                                    t.ciudad as ciudad,
                                    t.cp as cp,
                                    t.provincia as provincia,
                                    t.pais as pais
                            FROM
                                    negocio.mdp_personas mp
                                    INNER JOIN negocio.sga_alumnos sa ON mp.persona = sa.persona
                                    INNER JOIN negocio.sga_propuestas pr ON sa.propuesta = pr.propuesta
                                    LEFT JOIN LATERAL negocio_pers.sp_domicilio_persona(mp.persona, 'PROC') as t ON TRUE
                            {where_clause_str}
                            order by nombre_completo
        """
        # Combina los parámetros del CASE con los parámetros de la cláusula WHERE
        final_params = case_params + sql_params
        cursor.execute(sql_query, final_params) # Pasa todos los parámetros a la ejecución de la consulta
        columns = [col[0] for col in cursor.description]
        cursadas_data = cursor.fetchall()

    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cursadas"

    # Escribir los encabezados (columnas de la consulta SQL)
    ws.append(columns)

    # Escribir los datos
    for row in cursadas_data:
        processed_row = []
        for cell in row:
            if isinstance(cell, (date, datetime)):
                processed_row.append(cell.strftime('%Y-%m-%d'))
            else:
                processed_row.append(cell)
        ws.append(processed_row)

    # Configurar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_cursadas.xlsx"' # Cambiado el nombre del archivo
    wb.save(response) # Guarda el libro de trabajo en la respuesta HTTP
    return response


############################################################################
#PROMEDIO HISTORICO EN EXCEL
#############################################################################
@login_required
def export_promedio_historico_excel(request):
    # La lógica para obtener los datos es la misma que en promedio_historico_report
    # pero sin paginación, ya que queremos exportar TODOS los datos.

    all_promedios_data = []
    with connection.cursor() as cursor:
        sql_query = """
            SELECT ph.propuesta AS codigo_propuesta,
                   sp.nombre AS nombre_propuesta,
                   EXTRACT(YEAR FROM ph.fecha_generacion) AS anio,
                   ph.fecha_generacion AS generacion_raw,
                   ph.egresados_fecha_desde AS desde_raw,
                   ph.egresados_fecha_hasta AS hasta_raw,
                   ph.promedio_sin_aplazos AS sin_aplazos,
                   ph.promedio_general AS promedio,
                   ph.promedio_general_ponderado AS promedio_ponderado,
                   ph.desviacion_std AS desviacion,
                   ph.cant_egresados AS cantidad_egresados,
                   ph.cant_mat_total AS total_materias,
                   ph.cant_aprob AS materias_aprobadas
            FROM negocio.sga_propuestas sp,
                 negocio_pers.sga835_promedios_historicos ph
            WHERE sp.propuesta = ph.propuesta
            ORDER BY 3 DESC, 4 DESC, 1
        """
        cursor.execute(sql_query)

        columns = [col[0] for col in cursor.description]
        raw_data = cursor.fetchall()

        # Procesar los datos de la misma manera que en la vista principal
        for row in raw_data:
            row_dict = dict(zip(columns, row))

            # Convertir a objetos date/datetime explícitamente si es necesario
            if isinstance(row_dict['desde_raw'], datetime):
                row_dict['desde'] = row_dict['desde_raw'].date()
            else:
                row_dict['desde'] = row_dict['desde_raw']

            if isinstance(row_dict['hasta_raw'], datetime):
                row_dict['hasta'] = row_dict['hasta_raw'].date()
            else:
                row_dict['hasta'] = row_dict['hasta_raw']

            row_dict['generacion'] = row_dict['generacion_raw']

            all_promedios_data.append(row_dict)

    # Crear un libro de Excel y una hoja
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Promedios Históricos"

    # Encabezados de la tabla Excel
    headers = [
        "Código Propuesta", "Nombre Propuesta", "Año", "Fecha Generación",
        "Egresados Desde", "Egresados Hasta", "Promedio Sin Aplazos",
        "Promedio General", "Promedio Ponderado", "Desviación Std",
        "Cantidad Egresados", "Total Materias", "Materias Aprobadas"
    ]
    sheet.append(headers)

    # Añadir los datos
    for item in all_promedios_data:
        row_values = [
            item.get('codigo_propuesta'),
            item.get('nombre_propuesta'),
            item.get('anio'),
            item.get('generacion'), # Ya es un datetime object, openpyxl lo maneja bien
            item.get('desde'),      # Ya es un date object, openpyxl lo maneja bien
            item.get('hasta'),      # Ya es un date object, openpyxl lo maneja bien
            item.get('sin_aplazos'),
            item.get('promedio'),
            item.get('promedio_ponderado'),
            item.get('desviacion'),
            item.get('cantidad_egresados'),
            item.get('total_materias'),
            item.get('materias_aprobadas'),
        ]
        sheet.append(row_values)

    # Configurar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=promedios_historicos.xlsx'

    # Guardar el libro en la respuesta HTTP
    workbook.save(response)

    return response


############################################################################
#RESULTADO DE CURSADAS EN EXCEL
#############################################################################

@login_required
def export_resultado_cursadas_excel(request):
    # Inicializa el formulario con los parámetros GET recibidos
    form = ResultadoCursaFilterForm(request.GET)

    # Valores predeterminados (por si acaso no se aplican filtros o son inválidos)
    anio_filter = None
    propuesta_ids_filter = []

    # Validar los datos del formulario (los mismos filtros que en la vista del reporte)
    if form.is_valid():
        anio_filter = form.cleaned_data.get('anio')
        propuesta_ids_filter = form.cleaned_data.get('propuesta_ids')

    # Construir la consulta SQL dinámicamente basada en los filtros
    sql_where_clauses = ["ac.estado = 'C'"]
    sql_params = []

    if anio_filter:
        sql_where_clauses.append("c.anio_academico = %s ")
        sql_params.append(anio_filter)

    if propuesta_ids_filter:
        # Crea una cadena de marcadores de posición (%s) para la cláusula IN
        placeholders = ','.join(['%s'] * len(propuesta_ids_filter))
        sql_where_clauses.append(f"negocio_pers.get_propuesta_de_actividad(c.elemento) IN ({placeholders}) ")
        # Asegúrate de que los IDs sean enteros si tu base de datos los espera así
        sql_params.extend([int(p_id) for p_id in propuesta_ids_filter])

    # Unir las cláusulas WHERE
    where_clause_str = " AND ".join(sql_where_clauses)
    if where_clause_str: # Solo añade WHERE si hay cláusulas
        where_clause_str = "WHERE " + where_clause_str

    with connection.cursor() as cursor:
        sql_query = f"""
                 SET search_path TO negocio, negocio_pers;

                    SELECT
                            c.anio_academico AS anio,
                            c.comision_nombre AS nombre_comision,
                            c.comision AS comision,
                            c.elemento_nombre AS materia,
                            c.elemento_codigo AS codigo_materia,
                            CASE
                                WHEN ac.origen = 'R' THEN 'Acta_Regular'
                                WHEN ac.origen = 'P' THEN 'Acta_Promocion'
                            ELSE 'Acta_No_Valida'
                            END AS tipo_acta,
                        SUM(CASE WHEN negocio_pers.f_es_recursante(ad.alumno, c.anio_academico::integer, c.elemento, ic.fecha_inscripcion::date) = 'S' THEN 1 ELSE 0 END) AS recursantes_total,
                        SUM(CASE WHEN negocio_pers.f_es_recursante(ad.alumno, c.anio_academico::integer, c.elemento, ic.fecha_inscripcion::date) = 'N' THEN 1 ELSE 0 END) AS no_recursantes_total,
                        SUM(CASE WHEN negocio_pers.f_es_recursante(ad.alumno, c.anio_academico::integer, c.elemento, ic.fecha_inscripcion::date) = 'S' AND ad.resultado = 'A' THEN 1 ELSE 0 END) AS recursantes_aprobados,
                        SUM(CASE WHEN negocio_pers.f_es_recursante(ad.alumno, c.anio_academico::integer, c.elemento, ic.fecha_inscripcion::date) = 'S' AND ad.resultado != 'A' THEN 1 ELSE 0 END) AS recursantes_no_aprobados,
                        SUM(CASE WHEN negocio_pers.f_es_recursante(ad.alumno, c.anio_academico::integer, c.elemento, ic.fecha_inscripcion::date) = 'N' AND ad.resultado = 'A' THEN 1 ELSE 0 END) AS no_recursantes_aprobados,
                        SUM(CASE WHEN negocio_pers.f_es_recursante(ad.alumno, c.anio_academico::integer, c.elemento, ic.fecha_inscripcion::date) = 'N' AND ad.resultado != 'A' THEN 1 ELSE 0 END) AS no_recursantes_no_aprobados,
                        negocio_pers.get_docentes_de_una_comision(c.comision) as docentes
                    FROM
                        negocio.vw_comisiones c
                    INNER JOIN
                        negocio.sga_actas ac ON c.comision = ac.comision
                    INNER JOIN
                        negocio.sga_actas_detalle ad ON ac.id_acta = ad.id_acta
                    INNER JOIN
                        negocio.sga_insc_cursada ic ON c.comision = ic.comision AND ad.alumno = ic.alumno
                    INNER JOIN
                        negocio.sga_alumnos a ON ad.alumno = a.alumno -- Aunque 'a.legajo' no se usa en el SELECT final, mantenemos el JOIN por si hay alguna dependencia implícita.
                    {where_clause_str}
                    GROUP BY
                        c.anio_academico,
                        c.comision_nombre,
                        c.comision,
                        c.elemento_nombre,
                        c.elemento_codigo,
                        CASE
                            WHEN ac.origen = 'R' THEN 'Acta_Regular'
                            WHEN ac.origen = 'P' THEN 'Acta_Promocion'
                            ELSE 'Acta_No_Valida'
                        END
                    ORDER BY
                        c.anio_academico, c.comision_nombre, tipo_acta
        """
        cursor.execute(sql_query, sql_params) # Pasa los parámetros a la ejecución de la consulta
        columns = [col[0] for col in cursor.description]
        resultado_cursadas_data = cursor.fetchall()

    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultado_de_Cursadas"

    # Escribir los encabezados (columnas de la consulta SQL)
    ws.append(columns)

    # Escribir los datos
    for row in resultado_cursadas_data:
        processed_row = []
        for cell in row:
            if isinstance(cell, (date, datetime)):
                processed_row.append(cell.strftime('%Y-%m-%d'))
            else:
                processed_row.append(cell)
        ws.append(processed_row)

    # Configurar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_resultado_cursadas.xlsx"'
    wb.save(response) # Guarda el libro de trabajo en la respuesta HTTP
    return response # No se retorna una tupla, solo el objeto HttpResponse


############################################################################
#RESULTADO DE INGRESANTES POR AÑO EN EXCEL
#############################################################################

@login_required
def export_ingresantes_excel(request):
 # Inicializa el formulario con los parámetros GET recibidos
    form = IngresantesFilterForm(request.GET)

    # Valores predeterminados (por si acaso no se aplican filtros o son inválidos)
    anio_filter = None
    propuesta_ids_filter = []

    # Validar los datos del formulario (los mismos filtros que en la vista del reporte)
    if form.is_valid():
        anio_filter = form.cleaned_data.get('anio')
        propuesta_ids_filter = form.cleaned_data.get('propuesta_ids')

    # Construir la consulta SQL dinámicamente basada en los filtros
    sql_where_clauses = ["pv.estado = 'V'"]
    sql_params = []

    if anio_filter:
        sql_where_clauses.append("negocio_pers.anioingreso(a.alumno) = %s ")
        sql_params.append(anio_filter)

    if propuesta_ids_filter:
        # Crea una cadena de marcadores de posición (%s) para la cláusula IN
        placeholders = ','.join(['%s'] * len(propuesta_ids_filter))
        sql_where_clauses.append(f"sp.propuesta  IN ({placeholders}) ")
        # Asegúrate de que los IDs sean enteros si tu base de datos los espera así
        sql_params.extend([int(p_id) for p_id in propuesta_ids_filter])

    # Unir las cláusulas WHERE
    where_clause_str = " AND ".join(sql_where_clauses)
    if where_clause_str: # Solo añade WHERE si hay cláusulas
        where_clause_str = "WHERE " + where_clause_str

    with connection.cursor() as cursor:
        sql_query = f"""
                 SET search_path TO negocio, negocio_pers;

                    SELECT
                            p.nombre AS plan_estudios,
                            COUNT(a.alumno) AS cantidad_alumnos,
                            negocio_pers.anioingreso(a.alumno) AS anio_ingreso,
                            sp.nombre AS carrera
                        FROM
                            negocio.sga_alumnos a
                        INNER JOIN
                            negocio.sga_propuestas sp ON a.propuesta = sp.propuesta
                        INNER JOIN
                            negocio.sga_planes_versiones pv ON a.plan_version = pv.plan_version
                        INNER JOIN
                            negocio.sga_planes p ON pv.plan = p.plan AND sp.propuesta = p.propuesta
                        {where_clause_str}
                        GROUP BY
                            p.nombre,
                            negocio_pers.anioingreso(a.alumno),
                            sp.nombre
                        ORDER BY
                            p.nombre,
                            negocio_pers.anioingreso(a.alumno) ASC;
        """
        cursor.execute(sql_query, sql_params) # Pasa los parámetros a la ejecución de la consulta
        columns = [col[0] for col in cursor.description]
        ingresantes_data = cursor.fetchall()

    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ingresantes"

    # Escribir los encabezados (columnas de la consulta SQL)
    ws.append(columns)

    # Escribir los datos
    for row in ingresantes_data:
        processed_row = []
        for cell in row:
            if isinstance(cell, (date, datetime)):
                processed_row.append(cell.strftime('%Y-%m-%d'))
            else:
                processed_row.append(cell)
        ws.append(processed_row)

    # Configurar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_ingresantes.xlsx"'
    wb.save(response) # Guarda el libro de trabajo en la respuesta HTTP
    return response # No se retorna una tupla, solo el objeto HttpResponse

#########################################################################
# DOCENTES POR CARRERA O DEPARTAMENTO EN EXCEL
############################################################################

@login_required
def export_docentes_cardpto_excel(request):
    form = DocentesFilterForm(request.GET)

    results = []
    anio_filter = None

    # Listas para almacenar las cadenas SQL de cada rama y sus parámetros individuales
    union_query_strings = []
    union_query_params_individual = [] # Para acumular los parámetros de cada rama

    if request.GET:
        if form.is_valid():
            report_generated = True

            anio_filter = form.cleaned_data.get('anio')
            propuesta_ids_filter = form.cleaned_data.get('propuesta_ids') # .cleaned_data es correcto
            dptos_ids_filter = form.cleaned_data.get('dptos_ids')

            if propuesta_ids_filter:
                propuesta_ids_filter = [int(p_id) for p_id in propuesta_ids_filter if p_id]
            else:
                propuesta_ids_filter = []

            if dptos_ids_filter:
                dptos_ids_filter = [int(d_id) for d_id in dptos_ids_filter if d_id]
            else:
                dptos_ids_filter = []

            try:
                with connection.cursor() as cursor:
                    # --- Construcción Dinámica de las Ramas SQL y sus Parámetros ---

                    # Lógica para la Rama de Propuestas
                    # Incluir esta rama si:
                    # 1. Hay filtros de propuesta seleccionados, O
                    # 2. NO hay filtros de propuesta Y NO hay filtros de departamento (es decir, mostrar ambas por defecto)
                    if propuesta_ids_filter or (not propuesta_ids_filter and not dptos_ids_filter):
                        propuesta_branch_where_clauses = ["c.estado = 'A'"]
                        propuesta_branch_params = []

                        if anio_filter:
                            propuesta_branch_where_clauses.append("c.anio_academico = %s")
                            propuesta_branch_params.append(anio_filter)

                        if propuesta_ids_filter:
                            propuesta_branch_where_clauses.append(f"pro.propuesta IN ({','.join(['%s'] * len(propuesta_ids_filter))})")
                            propuesta_branch_params.extend(propuesta_ids_filter)

                        union_query_strings.append(f"""
                            SELECT
                                'PROPUESTA' AS tipo_filtro,
                                pro.nombre AS carrera,
                                pro.codigo AS codigo_carrera,
                                COUNT(DISTINCT dc.docente) AS cantidad_docentes_unicos_por_carrera,
                                negocio_pers.get_cant_docentesduplicados(c.anio_academico::integer, pro.propuesta, NULL) AS cantidad_docentes_por_carrera
                            FROM
                                negocio.vw_comisiones AS c
                            INNER JOIN negocio.sga_docentes_comision AS dc ON c.comision = dc.comision
                            INNER JOIN negocio.sga_docentes AS d ON dc.docente = d.docente
                            INNER JOIN negocio.mdp_personas AS p ON d.persona = p.persona
                            INNER JOIN negocio.sga_elementos AS e ON c.elemento = e.elemento
                            INNER JOIN negocio.sga_elementos_revision AS er ON e.elemento = er.elemento
                            INNER JOIN negocio.sga_elementos_plan AS ep ON er.elemento_revision = ep.elemento_revision
                            INNER JOIN negocio.sga_planes_versiones AS pv ON ep.plan_version = pv.plan_version
                            INNER JOIN negocio.sga_planes AS pr ON pv.plan = pr.plan
                            INNER JOIN negocio.sga_propuestas AS pro ON pr.propuesta = pro.propuesta
                            INNER JOIN negocio.sga_elementos_ra AS era ON e.elemento = era.elemento
                            INNER JOIN negocio.sga_responsables_academicas AS ra ON ra.responsable_academica = era.responsable_academica AND ra.responsable_academica_tipo = 2
                            WHERE
                                {' AND '.join(propuesta_branch_where_clauses)}
                            GROUP BY
                                tipo_filtro,
                                carrera,
                                codigo_carrera,
                                c.anio_academico,
                                pro.propuesta
                        """)
                        union_query_params_individual.extend(propuesta_branch_params)


                    # Lógica para la Rama de Departamentos (Responsables Académicos)
                    # Incluir esta rama si:
                    # 1. Hay filtros de departamento seleccionados, O
                    # 2. NO hay filtros de propuesta Y NO hay filtros de departamento (es decir, mostrar ambas por defecto)
                    if dptos_ids_filter or (not propuesta_ids_filter and not dptos_ids_filter):
                        dpto_branch_where_clauses = ["c.estado = 'A'"]
                        dpto_branch_params = []

                        if anio_filter:
                            dpto_branch_where_clauses.append("c.anio_academico = %s")
                            dpto_branch_params.append(anio_filter)

                        if dptos_ids_filter:
                            dpto_branch_where_clauses.append(f"ra.responsable_academica IN ({','.join(['%s'] * len(dptos_ids_filter))})")
                            dpto_branch_params.extend(dptos_ids_filter)

                        union_query_strings.append(f"""
                            SELECT
                                'DEPARTAMENTO' AS tipo_filtro,
                                ra.nombre AS carrera,
                                ra.codigo AS codigo_carrera,
                                COUNT(DISTINCT dc.docente) AS cantidad_docentes_unicos_por_carrera,
                                negocio_pers.get_cant_docentesduplicados(c.anio_academico::integer, NULL, ra.responsable_academica) AS cantidad_docentes_por_carrera
                            FROM
                                negocio.vw_comisiones AS c
                            INNER JOIN negocio.sga_docentes_comision AS dc ON c.comision = dc.comision
                            INNER JOIN negocio.sga_docentes AS d ON dc.docente = d.docente
                            INNER JOIN negocio.mdp_personas AS p ON d.persona = p.persona
                            INNER JOIN negocio.sga_elementos AS e ON c.elemento = e.elemento
                            INNER JOIN negocio.sga_elementos_revision AS er ON e.elemento = er.elemento
                            INNER JOIN negocio.sga_elementos_plan AS ep ON er.elemento_revision = ep.elemento_revision
                            INNER JOIN negocio.sga_planes_versiones AS pv ON ep.plan_version = pv.plan_version
                            INNER JOIN negocio.sga_planes AS pr ON pv.plan = pr.plan
                            INNER JOIN negocio.sga_propuestas AS pro ON pr.propuesta = pro.propuesta
                            INNER JOIN negocio.sga_elementos_ra AS era ON e.elemento = era.elemento
                            INNER JOIN negocio.sga_responsables_academicas AS ra ON ra.responsable_academica = era.responsable_academica AND ra.responsable_academica_tipo = 2
                            WHERE
                                {' AND '.join(dpto_branch_where_clauses)}
                            GROUP BY
                                tipo_filtro,
                                carrera,
                                codigo_carrera,
                                c.anio_academico,
                                ra.responsable_academica
                        """)
                        union_query_params_individual.extend(dpto_branch_params)


                    # --- Ejecución de la Consulta ---
                    if not union_query_strings:
                        # Si ninguna rama fue seleccionada (ej. si los filtros específicos no están y no se desea el default)
                        report_generated = False
                        # Opcional: form.add_error(None, "Por favor, selecciona al menos un filtro de Carrera o Departamento.")
                    else:
                        final_sql_query = " UNION ALL ".join(union_query_strings) + " ORDER BY tipo_filtro, carrera, codigo_carrera;"

                        # Los parámetros finales son la concatenación de todos los parámetros individuales
                        final_sql_params = union_query_params_individual

                        cursor.execute(final_sql_query, final_sql_params)
                        columns = [col[0] for col in cursor.description]
                        results = cursor.fetchall()
                        # Crear un nuevo libro de Excel
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "Docentes por Carrera o Departamento"

                        # Escribir los encabezados (columnas de la consulta SQL)
                        ws.append(columns)

                        # Escribir los datos
                        for row in results:
                            processed_row = []
                            for cell in row:
                                if isinstance(cell, (date, datetime)):
                                    processed_row.append(cell.strftime('%Y-%m-%d'))
                                else:
                                    processed_row.append(cell)
                            ws.append(processed_row)

                            # Configurar la respuesta HTTP para la descarga del archivo
                            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                            response['Content-Disposition'] = 'attachment; filename="docentes_carrera_dpto.xlsx"'
                            wb.save(response) # Guarda el libro de trabajo en la respuesta HTTP

            except psycopg2.Error as e:
                print(f"Error en la consulta SQL de docentes_x_carrera_dpto_view: {e}")
                results = []
                form.add_error(None, f"No se pudo cargar el reporte debido a un error en la base de datos: {e}")
            except Exception as e:
                print(f"Error inesperado en docentes_x_carrera_dpto_view: {e}")
                results = []
                form.add_error(None, f"No se pudo cargar el reporte. Error: {e}")
        else:
            pass # El formulario no es válido, los errores ya están en form.errors
    else:
        pass # No hay parámetros GET, es la primera carga de la página

    return response # No se retorna una tupla, solo el objeto HttpResponse


#########################################################################
# ALUMNOS RETENIDOS 1ER AÑO EN EXCEL
############################################################################

@login_required
def export_retenidos_por_carrera_view(request):
    form = IngresantesFilterForm(request.GET)

    labels = []
    data_ingresantes = []
    data_retenidos = []
    processed_columns = []
    results = []
    anio_filter = None
    report_generated = False

    if request.GET:
        if form.is_valid():
            report_generated = True

            anio_filter = form.cleaned_data.get('anio')
            propuesta_ids_filter = form.cleaned_data.get('propuesta_ids') # Corregido a .cleaned_data

            # Asegúrate de que anio_filter sea un entero para la suma
            if anio_filter:
                try:
                    anio_filter = int(anio_filter)
                    anio_siguiente = anio_filter + 1 # <--- ¡CALCULA EL AÑO SIGUIENTE AQUÍ!
                except (ValueError, TypeError):
                    anio_filter = None
                    anio_siguiente = None # Si el año no es válido, tampoco lo será el siguiente
            else:
                anio_siguiente = None # Si no hay anio_filter, no hay anio_siguiente

            propuesta_ids_filter = [int(p_id) for p_id in propuesta_ids_filter if p_id]

            sql_where_clauses = [] # Comienza vacía, no con ""
            sql_params = []

            if anio_filter:
                sql_where_clauses.append("negocio_pers.anioingreso(sa.alumno) = %s")
                sql_params.append(anio_filter)

            if propuesta_ids_filter:
                placeholders = ','.join(['%s'] * len(propuesta_ids_filter))
                sql_where_clauses.append(f"pr.propuesta IN ({placeholders})")
                sql_params.extend(propuesta_ids_filter)

            where_clause_str = ""
            if sql_where_clauses: # Solo agrega WHERE si hay cláusulas
                where_clause_str = "WHERE " + " AND ".join(sql_where_clauses)

            if anio_filter is None:
                report_generated = False
                form.add_error(None, "Por favor, selecciona un año válido para generar el reporte.")
                # Saltamos el resto de la lógica de la base de datos
                context = {
                    'form': form,
                    'anio_ingreso_filtrado': 'Todos',
                    'labels': [],
                    'data_unicos': [],
                    'data_totales': [],
                    'report_data': [],
                    'columns': [],
                    'report_generated': False,
                }
                return render(request, 'account/estadisticas/retenidos/retenidos_report.html', context)


            try:
                with connection.cursor() as cursor:
                    sql_query = f"""
                        SELECT
                            pr.nombre AS nombre_carrera,
                            pr.codigo AS codigo_carrera,
                            COUNT(sa.legajo) AS total_ingresantes,
                            COUNT(sa.legajo) FILTER (
                                WHERE
                                    EXISTS (
                                        SELECT 1
                                        FROM
                                            negocio.vw_comisiones c
                                        INNER JOIN negocio.sga_insc_cursada ic ON c.comision = ic.comision
                                        WHERE
                                            ic.alumno = sa.alumno
                                            AND c.anio_academico = %s
                                            AND ic.estado = 'A'
                                    )
                            ) AS alumnos_retenidos
                        FROM
                            negocio.mdp_personas mp
                        INNER JOIN negocio.sga_alumnos sa ON mp.persona = sa.persona
                        INNER JOIN negocio.sga_propuestas pr ON sa.propuesta = pr.propuesta
                        {where_clause_str}
                        GROUP BY
                            pr.nombre,
                            pr.codigo
                        ORDER BY
                            pr.nombre;
                    """
                    # Asegúrate de que anio_siguiente esté en sql_params
                    # Lo ideal es que vaya al principio si es el primer %s
                    # de la subconsulta del FILTER.
                    # El orden de los parámetros es CRUCIAL.

                    # Para la subconsulta, anio_academico es el primer %s en el SELECT.
                    # Luego vienen los %s de anioingreso y los %s de pr.propuesta IN (...).
                    # Por lo tanto, anio_siguiente debe ser el PRIMER parámetro en la lista.

                    # Vamos a crear la lista de parámetros final de forma explícita:
                    final_sql_params = [anio_siguiente] # El primer parámetro para el c.anio_academico del FILTER
                    final_sql_params.extend(sql_params) # Luego, los parámetros del WHERE principal (anio_filter, propuesta_ids)

                    cursor.execute(sql_query, final_sql_params)
                    columns = [col[0] for col in cursor.description]
                    results = cursor.fetchall()
                # Crear un nuevo libro de Excel
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Alumnos Retenidos de Primer Año"

                # Escribir los encabezados (columnas de la consulta SQL)
                    ws.append(columns)

                # Escribir los datos
                for row in results:
                    processed_row = []
                    for cell in row:
                        if isinstance(cell, (date, datetime)):
                                    processed_row.append(cell.strftime('%Y-%m-%d'))
                        else:
                                    processed_row.append(cell)
                    ws.append(processed_row)

                # Configurar la respuesta HTTP para la descarga del archivo
                    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename="alumnos_retenidos.xlsx"'
                    wb.save(response) # Guarda el libro de trabajo en la respuesta HTTP

            except psycopg2.Error as e: # Es bueno capturar errores específicos de la DB
                print(f"Error de base de datos en retenidos_por_carrera_view: {e}")
                results = []
                report_generated = False
                form.add_error(None, f"No se pudo cargar el reporte debido a un error en la base de datos: {e}")
            except Exception as e:
                print(f"Error inesperado en retenidos_por_carrera_view: {e}")
                results = []
                report_generated = False
                form.add_error(None, f"No se pudo cargar el reporte. Error: {e}")
        else:
            pass
    else:
        pass

    return response




###########################################################################################
### RANGO ETARIO REPORTE
############################################################################################

@login_required
def rango_etario_view(request):

    form = ResultadoCursaFilterForm(request.GET)

    labels = []  # Etiquetas para el eje X del gráfico (rangos etarios)
    male_counts = []  # Datos para la serie de hombres en el gráfico
    female_counts = []  # Datos para la serie de mujeres en el gráfico
    table_data = [] # Usamos table_data para los datos de la tabla
    chart_data_present = False  # Bandera para indicar si hay datos para el gráfico

     # Esta bandera indicará si se generó un reporte (después de aplicar filtros)
    report_generated = False

    # Solo procesamos si el formulario es válido y tiene datos de filtro
    if form.is_valid():
        anio_filter = form.cleaned_data.get('anio')
        propuesta_ids_filter = form.cleaned_data.get('propuesta_ids')
        report_generated = True


        # Ejecuta la consulta solo si se han seleccionado ambos filtros
        if anio_filter and propuesta_ids_filter:
            anio_filter = int(anio_filter)
            propuesta_ids_tuple = tuple(map(int, propuesta_ids_filter))

            with connection.cursor() as cursor:
                sql_query = f"""
                WITH alumnos_con_edad AS (
                    SELECT
                        mp.persona,
                        pr.codigo,
                        sa.legajo,
                        mp.sexo,
                        negocio_pers.f_edad_que_tenia_al(
                            mp.fecha_nacimiento,
                            MAKE_DATE(%s, EXTRACT(MONTH FROM NOW())::INT, EXTRACT(DAY FROM NOW())::INT)
                        ) AS edad_calculada,
                        sa.propuesta,
                        sa.alumno,
                        pr.nombre
                    FROM
                        negocio.mdp_personas mp
                    INNER JOIN negocio.sga_alumnos sa ON mp.persona = sa.persona
                    INNER JOIN negocio.sga_propuestas pr ON sa.propuesta = pr.propuesta
                    WHERE pr.propuesta IN %s
                ),
                alumnos_filtrados_comision AS (
                    SELECT
                        ace.*
                    FROM
                        alumnos_con_edad ace
                    WHERE
                        EXISTS (
                            SELECT 1
                            FROM
                                negocio.vw_comisiones c
                            INNER JOIN negocio.sga_insc_cursada ic ON c.comision = ic.comision
                            WHERE
                                ic.alumno = ace.alumno
                                AND c.anio_academico = %s
                                AND ic.estado = 'A'
                        )
                )
                SELECT
                    CASE
                        WHEN edad_calculada BETWEEN 17 AND 22 THEN '17-22'
                        WHEN edad_calculada BETWEEN 23 AND 27 THEN '23-27'
                        WHEN edad_calculada BETWEEN 28 AND 32 THEN '28-32'
                        WHEN edad_calculada BETWEEN 33 AND 37 THEN '33-37'
                        WHEN edad_calculada BETWEEN 38 AND 42 THEN '38-42'
                        WHEN edad_calculada >= 43 THEN 'Mas de 43 años'
                        ELSE 'Desconocido'
                    END AS rango_etario,
                    sexo,
                    nombre,
                    COUNT(*) AS total_alumnos
                FROM
                    alumnos_filtrados_comision
                GROUP BY
                    rango_etario, sexo, nombre
                ORDER BY
                    rango_etario, sexo, nombre;
                """
                cursor.execute(sql_query, [anio_filter, propuesta_ids_tuple, anio_filter])
                rows = cursor.fetchall()

            # --- Lógica de procesamiento de datos para Chart.js y la Tabla ---
            # Este diccionario consolidará los datos para cada rango etario, separando Hombres y Mujeres.
            chart_data = {}
            for row in rows:
                rango_etario = row[0]
                sexo = row[1]
                carrera = row[2]  # Carrera, aunque no se usa en el gráfico, puede ser útil para la tabla
                total_alumnos = row[3]

                if rango_etario not in chart_data:
                    chart_data[rango_etario] = {'M': 0, 'F': 0} # Inicializa con 0 para ambos sexos

                if sexo == 'M':
                    chart_data[rango_etario]['M'] += total_alumnos
                elif sexo == 'F':
                    chart_data[rango_etario]['F'] += total_alumnos

            # Solo si chart_data tiene datos, preparamos las listas y la tabla
            if chart_data:
                # Función para ordenar los rangos etarios
                def sort_key(item):
                    if item == 'Mas de 43 años':
                        return 1000
                    elif item == 'Desconocido':
                        return 2000
                    else:
                        try:
                            return int(item.split('-')[0])
                        except ValueError:
                            return 900

                # Generar las etiquetas para el gráfico (eje X), ordenadas
                labels = sorted(chart_data.keys(), key=sort_key)

                # Generar las listas de conteos para Hombres y Mujeres, en el mismo orden que las labels
                male_counts = [chart_data[label]['M'] for label in labels]
                female_counts = [chart_data[label]['F'] for label in labels]

                # Generar los datos para la tabla HTML a partir del chart_data consolidado
                for label in labels:
                    hombres = chart_data[label]['M']
                    mujeres = chart_data[label]['F']
                    total = hombres + mujeres
                    table_data.append({
                        'rango_etario': label,
                        'hombres': hombres,
                        'mujeres': mujeres,
                        'total': total,
                        'carrera': carrera,
                    })

                chart_data_present = True # Tenemos datos para mostrar
            else:
                chart_data_present = False # chart_data estaba vacío, no hay datos para mostrar
                report_generated = False
        else:
            chart_data_present = False # Faltan filtros (año o propuestas)
            report_generated = False

    else:
        chart_data_present = False # Formulario no válido o no enviado con filtros
        report_generated = False

    context = {
        'form': form,
        'labels': json.dumps(labels),
        'male_counts': json.dumps(male_counts),
        'female_counts': json.dumps(female_counts),
        'chart_data_present': chart_data_present,
        'table_data': table_data,
        'report_generated': report_generated, # ¡Pasar la bandera a la plantilla!

    }

    return render(request, 'account/estadisticas/etarios/rango_etario.html', context)


#########################################################################
# RANGO ETARIO EN EXCEL
############################################################################

def export_rango_etario_excel(request):
    form = ResultadoCursaFilterForm(request.GET)
    results = [] # Inicializamos results fuera del bloque para que esté siempre disponible

    # Solo si el formulario es válido y tiene los filtros necesarios
    if form.is_valid():
        anio_filter = form.cleaned_data.get('anio')
        propuesta_ids_filter = form.cleaned_data.get('propuesta_ids')

        if anio_filter and propuesta_ids_filter:
            try:
                anio_filter = int(anio_filter)
                propuesta_ids_tuple = tuple(map(int, propuesta_ids_filter))

                with connection.cursor() as cursor:

                    sql_query = f"""
                    WITH alumnos_con_edad AS (
                        SELECT
                            mp.persona,
                            pr.codigo,
                            sa.legajo,
                            mp.sexo,
                            negocio_pers.f_edad_que_tenia_al(
                                mp.fecha_nacimiento,
                                MAKE_DATE(%s, EXTRACT(MONTH FROM NOW())::INT, EXTRACT(DAY FROM NOW())::INT)
                            ) AS edad_calculada,
                            sa.propuesta,
                            sa.alumno
                        FROM
                            negocio.mdp_personas mp
                        INNER JOIN negocio.sga_alumnos sa ON mp.persona = sa.persona
                        INNER JOIN negocio.sga_propuestas pr ON sa.propuesta = pr.propuesta
                        WHERE pr.propuesta IN %s
                    ),
                    alumnos_filtrados_comision AS (
                        SELECT
                            ace.*
                        FROM
                            alumnos_con_edad ace
                        WHERE
                            EXISTS (
                                SELECT 1
                                FROM
                                    negocio.vw_comisiones c
                                INNER JOIN negocio.sga_insc_cursada ic ON c.comision = ic.comision
                                WHERE
                                    ic.alumno = ace.alumno
                                    AND c.anio_academico = %s
                                    AND ic.estado = 'A'
                            )
                    )
                    SELECT
                        T2.rango_etario,
                        T2.sexo,
                        T2.total_alumnos
                    FROM (
                        SELECT
                            CASE
                                WHEN edad_calculada BETWEEN 17 AND 22 THEN '17-22'
                                WHEN edad_calculada BETWEEN 23 AND 27 THEN '23-27'
                                WHEN edad_calculada BETWEEN 28 AND 32 THEN '28-32'
                                WHEN edad_calculada BETWEEN 33 AND 37 THEN '33-37'
                                WHEN edad_calculada BETWEEN 38 AND 42 THEN '38-42'
                                WHEN edad_calculada >= 43 THEN 'Mas de 43 años'
                                ELSE 'Desconocido'
                            END AS rango_etario,
                            sexo,
                            COUNT(*) AS total_alumnos
                        FROM
                            alumnos_filtrados_comision
                        GROUP BY
                            rango_etario, sexo
                        ORDER BY
                            rango_etario, sexo
                    ) AS T2;
                    """
                    cursor.execute(sql_query, [anio_filter, propuesta_ids_tuple, anio_filter])
                    # Obtenemos los nombres de las columnas para los encabezados del Excel
                    columns = [col[0] for col in cursor.description]
                    results = cursor.fetchall() # Capturamos todos los resultados

                # --- Generación del archivo Excel ---
                wb = openpyxl.Workbook()
                ws = wb.active
                # El título de la hoja puede incluir el año y las propuestas si lo deseas
                ws.title = f"RangoEtario_{anio_filter}"

                # Escribir los encabezados (columnas de la consulta SQL)
                ws.append(columns)

                # Escribir los datos
                for row in results:
                    processed_row = []
                    for cell in row:
                        if isinstance(cell, (date, datetime)):
                            # Formatear fechas si las hubiera en los resultados finales
                            processed_row.append(cell.strftime('%Y-%m-%d'))
                        else:
                            processed_row.append(cell)
                    ws.append(processed_row)

                # Configurar la respuesta HTTP para la descarga del archivo
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                filename = f"rango_etario_{anio_filter}"
                # Construir el nombre del archivo con los IDs de propuesta si hay muchos
                if len(propuesta_ids_filter) == 1:
                    filename += f"_propuesta_{propuesta_ids_filter[0]}"
                elif len(propuesta_ids_filter) > 1:
                    # Evita nombres de archivo demasiado largos si hay muchas propuestas
                    filename += "_multiples_propuestas"

                response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

                wb.save(response) # Guarda el libro de trabajo en la respuesta HTTP
                return response # Retorna la respuesta HTTP con el archivo

            except Exception as e:
                # Si ocurre algún error (DB o Python), maneja la excepción
                # y devuelve una respuesta HTTP que informe del error al usuario.
                # Considera loggear el error para debugging.
                print(f"Error al generar el archivo Excel: {e}")
                # Puedes mostrar un mensaje de error simple o redirigir a una página de error
                # Por simplicidad, devolveremos una respuesta HTTP 500 (Internal Server Error)
                return HttpResponse(f"Error al generar el reporte Excel: {e}", status=500)
    else:
        # Si el formulario no es válido o no se enviaron los filtros necesarios
        # puedes redirigir al usuario o mostrar un mensaje de error.
        # Aquí, por ejemplo, redirigimos a la vista del reporte HTML.
        from django.shortcuts import redirect
        return redirect('rango_etario_view') # Redirige a la vista del reporte HTML original

    # Si por alguna razón la lógica no retorna en los bloques anteriores (ej. no hay datos,
    # pero no se lanzó una excepción), asegúrate de tener un retorno por defecto.
    # Esto no debería ejecutarse si la lógica es completa.
    return HttpResponse("No se pudo generar el reporte Excel.", status=400)




###########################################################################################
### EGRESADOS POR AÑO DE INGRESO
############################################################################################

@login_required
def egresados_x_anio_view(request):

    form = EgresadosxAnioFilterForm(request.GET)
    egresados = []
    egresados_page_obj = None
    report_executed = False

    # Solo procesamos si el formulario es válido y tiene datos de filtro
    if form.is_valid():
        anio_filter = form.cleaned_data.get('anio')
        propuesta_ids_filter = form.cleaned_data.get('propuesta_ids')

        # Ejecuta la consulta solo si se han seleccionado ambos filtros
        if anio_filter and propuesta_ids_filter:
            anio_filter = int(anio_filter)
            propuesta_ids_str = ','.join(map(str, [int(p_id) for p_id in propuesta_ids_filter]))

            report_executed = True

            with connection.cursor() as cursor:

                sql_query = f"""
                    SELECT
                        negocio_pers.anioingreso(a.alumno) AS anioingreso,
                        EXTRACT(YEAR FROM sco.fecha_egreso) AS anio_egreso,
                        vp.sexo AS sexo,
                        a.plan_codigo AS plan_codigo,
                        a.propuesta_nombre AS carrera_nombre,
                        COUNT(a.alumno) AS cantidad_egresados
                    FROM
                        negocio.vw_alumnos a
                    INNER JOIN
                        negocio.sga_certificados_otorg sco ON a.alumno = sco.alumno AND a.persona = sco.persona
                    INNER JOIN
                        negocio.vw_personas vp ON a.persona = vp.persona
                    WHERE
                        a.propuesta IN ({propuesta_ids_str})
                        AND negocio_pers.anioingreso(a.alumno) >= %s
                        AND sco.fecha_egreso is not null

                    GROUP BY
                        negocio_pers.anioingreso(a.alumno),
                        EXTRACT(YEAR FROM sco.fecha_egreso),
                        vp.sexo,
                        a.plan_codigo,
                        a.propuesta_nombre
                    ORDER by
                        anioingreso asc,
                        anio_egreso ASC,
                        sexo ASC,
                        plan_codigo ASC,
                        cantidad_egresados DESC;
                """
                cursor.execute(sql_query, [anio_filter])
                columns = [col[0] for col in cursor.description]
                all_egresados_data = cursor.fetchall()

                egresados_list = []
                for row in all_egresados_data:
                        row_dict = dict(zip(columns, row))
                        egresados_list.append(row_dict)

                # El paginador y el formato de datos deben ocurrir después de obtener todos los datos,
                # pero pueden estar fuera del bloque del cursor ya que all_egresados_data ya está en memoria.
                paginator = Paginator(egresados_list, 30)
                page_number = request.GET.get('page')

                try:
                    egresados_page_obj = paginator.page(page_number)
                except PageNotAnInteger:
                    egresados_page_obj = paginator.page(1)
                except EmptyPage:
                    egresados_page_obj = paginator.page(paginator.num_pages)

                # Poblar la lista 'egresados' directamente desde egresados_page_obj.
                # Esto asegura que los datos estén correctamente estructurados para la plantilla
                # y evita posibles problemas de re-iteración con el cursor.
                for egresado_dict in egresados_page_obj:
                    egresados.append({
                        'anioingreso': egresado_dict['anioingreso'],
                        'anio_egreso': egresado_dict['anio_egreso'],
                        'sexo': egresado_dict['sexo'],
                        'plan_codigo': egresado_dict['plan_codigo'],
                        'carrera_nombre': egresado_dict['carrera_nombre'],
                        'cantidad_egresados': egresado_dict['cantidad_egresados'],
                    })

    context = {
        'form': form,
        'egresados': egresados,
        'egresados_page_obj': egresados_page_obj,
        'report_title': 'Reporte de Egresados',
        'report_executed': report_executed,
    }

    return render(request, 'account/estadisticas/egresados/egresados_x_anio.html', context)



#########################################################################
# DOCENTES POR  DEPARTAMENTO EN REPORTES
############################################################################

@login_required
def docentes_x_comision_report(request):
    form = DocentesFilterForm(request.GET)

    docentes = []
    docentes_page_obj = None
    report_executed = False

    if 'anio' in request.GET or 'propuesta_ids' in request.GET:
        if form.is_valid():
            anio_filter = form.cleaned_data.get('anio')
            dptos_ids_filter = form.cleaned_data.get('dptos_ids')

            if anio_filter and dptos_ids_filter:
                report_executed = True

            if dptos_ids_filter:
                dptos_ids_filter = ','.join(map(str, [int(p_id) for p_id in dptos_ids_filter]))

            else:
                dptos_ids_filter = []

            with connection.cursor() as cursor:
                    sql_query = f"""
                        select
                        ra.nombre AS carrera,
                        ra.codigo AS codigo_dpto,
                        COUNT(DISTINCT dc.docente) AS total_docentes,
                        e.nombre as nombre_materia,
                        c.comision_nombre as comision_nombre,
                        c.periodo_nombre as cursado,
                        negocio_pers.get_docentes_de_una_comision(c.comision) as nombre_docentes
                        FROM
                            negocio.vw_comisiones AS c
                        INNER JOIN
                            negocio.sga_docentes_comision AS dc ON c.comision = dc.comision
                        INNER JOIN
                            negocio.sga_docentes AS d ON dc.docente = d.docente
                        INNER JOIN
                            negocio.mdp_personas AS p ON d.persona = p.persona
                        INNER JOIN
                            negocio.sga_elementos AS e ON c.elemento = e.elemento
                        INNER JOIN
                            negocio.sga_elementos_revision AS er ON e.elemento = er.elemento
                        INNER JOIN
                            negocio.sga_elementos_plan AS ep ON er.elemento_revision = ep.elemento_revision
                        INNER JOIN
                            negocio.sga_planes_versiones AS pv ON ep.plan_version = pv.plan_version
                        INNER JOIN
                            negocio.sga_planes AS pr ON pv.plan = pr.plan
                        INNER JOIN
                            negocio.sga_propuestas AS pro ON pr.propuesta = pro.propuesta
                        INNER JOIN
                            negocio.sga_elementos_ra AS era ON e.elemento = era.elemento
                        INNER JOIN
                            negocio.sga_responsables_academicas AS ra ON ra.responsable_academica = era.responsable_academica AND ra.responsable_academica_tipo = 2
                        WHERE
                            c.anio_academico = %s
                            AND c.estado = 'A'
                            AND ra.responsable_academica IN ({dptos_ids_filter})
                        GROUP BY
                            carrera,
                            codigo_dpto,
                            comision_nombre,
                            c.periodo_nombre,
                            c.comision,
                            e.nombre,
                            c.anio_academico,
                            ra.responsable_academica
                        ORDER BY
                            nombre_materia,
                            codigo_dpto asc
                        """
                    cursor.execute(sql_query, [anio_filter])

                    columns = [col[0] for col in cursor.description]
                    all_docentes_data = cursor.fetchall()

                    docentes_list = []
                    for row in all_docentes_data:
                        row_dict = dict(zip(columns, row))
                        docentes_list.append(row_dict)

            # El paginador y el formato de datos deben ocurrir después de obtener todos los datos,
            # pero pueden estar fuera del bloque del cursor ya que all_docentes_data ya está en memoria.
            paginator = Paginator(docentes_list, 25)
            page_number = request.GET.get('page')

            try:
                    docentes_page_obj = paginator.page(page_number)
            except PageNotAnInteger:
                    docentes_page_obj = paginator.page(1)
            except EmptyPage:
                    docentes_page_obj = paginator.page(paginator.num_pages)

            # Poblar la lista 'docentes' directamente desde docentes_page_obj.
            # Esto asegura que los datos estén correctamente estructurados para la plantilla
            # y evita posibles problemas de re-iteración con el cursor.
            for docente_dict in docentes_page_obj:
                    docentes.append({
                        'carrera': docente_dict['carrera'],
                        'codigo_dpto': docente_dict['codigo_dpto'],
                        'total_docentes': docente_dict['total_docentes'],
                        'nombre_materia': docente_dict['nombre_materia'],
                        'comision_nombre': docente_dict['comision_nombre'],
                        'cursado': docente_dict['cursado'],
                        'nombre_docentes': docente_dict['nombre_docentes'],
                    })

    context = {
        'form': form,
        'docentes': docentes,
        'docentes_page_obj': docentes_page_obj,
        'report_title': 'Reporte de Comisiones por Docente',
        'report_executed': report_executed,
    }

    return render(request, 'account/reportes/docentes/docentes_x_comision_report.html', context)
